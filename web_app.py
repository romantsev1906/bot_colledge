from flask import Flask, render_template, request, jsonify, send_file, session
from flask_cors import CORS
import sqlite3
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from datetime import datetime
import os
import logging
from functools import wraps

app = Flask(__name__)
app.secret_key = os.urandom(24)
CORS(app)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Конфигурация БД - локальная SQLite (та же БД, что использует бот)
DB_PATH = 'students.db'  # БД находится в той же папке, что и скрипты

def get_db_connection():
    """Получение подключения к локальной SQLite БД"""
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn

def get_cursor(conn):
    """Получение курсора"""
    return conn.cursor()

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'Unauthorized'}), 401
        return f(*args, **kwargs)
    return decorated_function

def role_required(roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                return jsonify({'error': 'Unauthorized'}), 401
            if session.get('role') not in roles:
                return jsonify({'error': 'Forbidden'}), 403
            return f(*args, **kwargs)
        return decorated_function
    return decorator

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/auth', methods=['POST'])
def auth():
    """Авторизация по Telegram ID"""
    data = request.json
    user_id = data.get('user_id')
    
    if not user_id:
        return jsonify({'error': 'User ID required'}), 400
    
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    try:
        cursor.execute("SELECT user_id, role, full_name FROM users WHERE user_id = ?", (user_id,))
        user = cursor.fetchone()
        
        if user:
            session['user_id'] = user['user_id']
            session['role'] = user['role']
            session['full_name'] = user['full_name']
            conn.close()
            return jsonify({
                'success': True,
                'role': user['role'],
                'full_name': user['full_name']
            })
        else:
            conn.close()
            return jsonify({
                'success': False,
                'message': 'User not found. Please enter your ID manually.'
            }), 404
    except Exception as e:
        logger.error(f"Auth error: {e}")
        conn.close()
        return jsonify({
            'success': False,
            'message': f'Database error: {str(e)}'
        }), 500

@app.route('/api/auth/manual', methods=['POST'])
def auth_manual():
    """Ручная авторизация"""
    data = request.json
    user_id = int(data.get('user_id'))
    
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    try:
        cursor.execute("SELECT user_id, role, full_name FROM users WHERE user_id = ?", (user_id,))
        user = cursor.fetchone()
        
        if user:
            session['user_id'] = user['user_id']
            session['role'] = user['role']
            session['full_name'] = user['full_name']
            conn.close()
            return jsonify({
                'success': True,
                'role': user['role'],
                'full_name': user['full_name']
            })
        else:
            conn.close()
            return jsonify({'error': 'User not found'}), 404
    except Exception as e:
        logger.error(f"Auth manual error: {e}")
        conn.close()
        return jsonify({'error': f'Database error: {str(e)}'}), 500

@app.route('/api/user/info')
@login_required
def get_user_info():
    return jsonify({
        'user_id': session.get('user_id'),
        'role': session.get('role'),
        'full_name': session.get('full_name')
    })

@app.route('/api/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'success': True})

# ========== СТУДЕНТ ==========

@app.route('/student')
@login_required
@role_required(['student'])
def student_dashboard():
    return render_template('student/dashboard.html')

@app.route('/api/student/grades')
@login_required
@role_required(['student'])
def get_student_grades():
    student_id = session['user_id']
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    try:
        cursor.execute("""
            SELECT g.grade, g.date, d.name as discipline, u.full_name as teacher, 
                   COALESCE(k.description, 'Нет описания КТП') as description,
                   COALESCE(k.homework, 'Нет домашнего задания') as homework
            FROM grades g
            JOIN disciplines d ON g.discipline_id = d.discipline_id
            JOIN users u ON g.teacher_id = u.user_id
            LEFT JOIN ktp k ON g.ktp_id = k.ktp_id
            WHERE g.student_id = ?
            ORDER BY g.date DESC
        """, (student_id,))
        
        grades = [dict(row) for row in cursor.fetchall()]
        
        for grade in grades:
            if grade['grade'] == -1:
                grade['grade'] = 'н'
        
        conn.close()
        return jsonify(grades)
    except Exception as e:
        logger.error(f"Error getting student grades: {e}")
        conn.close()
        return jsonify({'error': f'Database error: {str(e)}'}), 500

@app.route('/api/student/teachers')
@login_required
@role_required(['student'])
def get_student_teachers():
    student_id = session['user_id']
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    try:
        cursor.execute("""
            SELECT t.teacher_id, u.full_name, st.tokens 
            FROM student_teacher st
            JOIN teachers t ON st.teacher_id = t.teacher_id
            JOIN users u ON t.teacher_id = u.user_id
            WHERE st.student_id = ?
        """, (student_id,))
        
        teachers = [dict(row) for row in cursor.fetchall()]
        conn.close()
        return jsonify(teachers)
    except Exception as e:
        logger.error(f"Error getting student teachers: {e}")
        conn.close()
        return jsonify({'error': f'Database error: {str(e)}'}), 500

@app.route('/api/student/tokens')
@login_required
@role_required(['student'])
def get_student_tokens():
    student_id = session['user_id']
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    cursor.execute("""
        SELECT t.teacher_id, u.full_name, st.tokens 
        FROM student_teacher st
        JOIN teachers t ON st.teacher_id = t.teacher_id
        JOIN users u ON t.teacher_id = u.user_id
        WHERE st.student_id = ?
    """, (student_id,))
    
    tokens = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(tokens)

@app.route('/api/student/rewards')
@login_required
@role_required(['student'])
def get_student_rewards():
    student_id = session['user_id']
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    cursor.execute("""
        SELECT r.reward_id, r.name, r.description, r.price, u.full_name as teacher_name, 
               d.name as discipline_name, r.teacher_id
        FROM rewards r
        JOIN disciplines d ON r.discipline_id = d.discipline_id
        JOIN student_teacher st ON r.teacher_id = st.teacher_id
        JOIN users u ON r.teacher_id = u.user_id
        WHERE st.student_id = ? AND r.is_enabled = 1
        ORDER BY d.name, r.name
    """, (student_id,))
    
    rewards = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(rewards)

@app.route('/api/student/rewards/buy', methods=['POST'])
@login_required
@role_required(['student'])
def buy_reward():
    student_id = session['user_id']
    data = request.json
    reward_id = data.get('reward_id')
    
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    cursor.execute("""
        SELECT r.price, r.name, r.teacher_id, d.name as discipline_name
        FROM rewards r
        JOIN disciplines d ON r.discipline_id = d.discipline_id
        WHERE r.reward_id = ?
    """, (reward_id,))
    reward = cursor.fetchone()
    
    if not reward:
        conn.close()
        return jsonify({'error': 'Reward not found'}), 404
    
    cursor.execute("SELECT tokens FROM student_teacher WHERE student_id = ? AND teacher_id = ?",
                  (student_id, reward['teacher_id']))
    balance = cursor.fetchone()
    
    if not balance or balance['tokens'] < reward['price']:
        conn.close()
        return jsonify({'error': 'Insufficient tokens'}), 400
    
    cursor.execute("""
        UPDATE student_teacher 
        SET tokens = tokens - ? 
        WHERE student_id = ? AND teacher_id = ?
    """, (reward['price'], student_id, reward['teacher_id']))
    
    cursor.execute("""
        INSERT INTO purchased_rewards (student_id, reward_id, teacher_id, purchase_date)
        VALUES (?, ?, ?, ?)
    """, (student_id, reward_id, reward['teacher_id'], datetime.now().strftime("%d-%m-%Y")))
    
    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'message': f'Награда "{reward["name"]}" куплена!'})

@app.route('/api/student/practices')
@login_required
@role_required(['student'])
def get_student_practices():
    student_id = session['user_id']
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    cursor.execute("""
        SELECT d.discipline_id, d.name, d.required_practices
        FROM disciplines d
        JOIN student_teacher st ON d.teacher_id = st.teacher_id
        WHERE st.student_id = ? AND d.group_name = (SELECT group_name FROM students WHERE student_id = ?)
    """, (student_id, student_id))
    disciplines = cursor.fetchall()
    
    result = []
    for disc in disciplines:
        cursor.execute("""
            SELECT COUNT(DISTINCT k.practice_number)
            FROM grades g
            JOIN ktp k ON g.ktp_id = k.ktp_id
            WHERE g.student_id = ? AND g.discipline_id = ? AND k.type = 'practice'
            AND g.grade >= 1 AND g.grade != -1
        """, (student_id, disc['discipline_id']))
        completed = cursor.fetchone()[0]
        
        result.append({
            'discipline_id': disc['discipline_id'],
            'name': disc['name'],
            'completed': completed,
            'required': disc['required_practices'] or 0,
            'status': 'Все практики сданы' if completed >= (disc['required_practices'] or 0) and (disc['required_practices'] or 0) > 0 else 'Есть несданные практики'
        })
    
    conn.close()
    return jsonify(result)

@app.route('/api/student/export/grades')
@login_required
@role_required(['student'])
def export_student_grades():
    """Экспорт оценок студента в Excel с логотипом"""
    student_id = session['user_id']
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    cursor.execute("""
        SELECT g.grade, g.date, d.name as discipline, u.full_name as teacher, 
               COALESCE(k.description, 'Нет описания КТП') as description,
               COALESCE(k.homework, 'Нет домашнего задания') as homework
        FROM grades g
        JOIN disciplines d ON g.discipline_id = d.discipline_id
        JOIN users u ON g.teacher_id = u.user_id
        LEFT JOIN ktp k ON g.ktp_id = k.ktp_id
        WHERE g.student_id = ?
        ORDER BY g.date DESC
    """, (student_id,))
    
    grades = cursor.fetchall()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Оценки"
    
    # Добавляем логотип
    logo_path = os.path.join('static', 'logo.png')
    if os.path.exists(logo_path):
        img = Image(logo_path)
        img.width = 100
        img.height = 100
        ws.add_image(img, 'A1')
    
    headers = ["Оценка", "Дата", "Дисциплина", "Преподаватель", "Описание КТП", "Домашнее задание"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for row_num, grade in enumerate(grades, 4):
        grade_val = "н" if grade['grade'] == -1 else str(grade['grade'])
        ws.cell(row=row_num, column=1, value=grade_val)
        ws.cell(row=row_num, column=2, value=grade['date'])
        ws.cell(row=row_num, column=3, value=grade['discipline'])
        ws.cell(row=row_num, column=4, value=grade['teacher'])
        ws.cell(row=row_num, column=5, value=grade['description'])
        ws.cell(row=row_num, column=6, value=grade['homework'])
    
    for col_num in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_num)
        max_length = max(
            len(str(ws.cell(row=row, column=col_num).value or "")) for row in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    filename = f"grades_{student_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join('temp', filename)
    os.makedirs('temp', exist_ok=True)
    wb.save(filepath)
    conn.close()
    
    return send_file(filepath, as_attachment=True, download_name=filename)

# ========== ПРЕПОДАВАТЕЛЬ ==========

@app.route('/teacher')
@login_required
@role_required(['teacher'])
def teacher_dashboard():
    return render_template('teacher/dashboard.html')

@app.route('/api/teacher/disciplines')
@login_required
@role_required(['teacher'])
def get_teacher_disciplines():
    teacher_id = session['user_id']
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    cursor.execute("SELECT discipline_id, name, group_name FROM disciplines WHERE teacher_id = ? ORDER BY name",
                  (teacher_id,))
    disciplines = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(disciplines)

@app.route('/api/teacher/groups')
@login_required
@role_required(['teacher'])
def get_teacher_groups():
    teacher_id = session['user_id']
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    cursor.execute("SELECT DISTINCT group_name FROM student_groups WHERE teacher_id = ? ORDER BY group_name",
                  (teacher_id,))
    groups = [row['group_name'] for row in cursor.fetchall()]
    conn.close()
    return jsonify(groups)

@app.route('/api/teacher/students')
@login_required
@role_required(['teacher'])
def get_teacher_students():
    teacher_id = session['user_id']
    group_name = request.args.get('group')
    
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    query = """
        SELECT DISTINCT s.student_id, u.full_name, s.group_name 
        FROM students s
        JOIN users u ON s.student_id = u.user_id
        JOIN student_teacher st ON s.student_id = st.student_id
        WHERE st.teacher_id = ?
    """
    params = [teacher_id]
    
    if group_name:
        query += " AND s.group_name = ?"
        params.append(group_name)
    
    query += " ORDER BY u.full_name"
    cursor.execute(query, params)
    students = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(students)

@app.route('/api/teacher/ktp')
@login_required
@role_required(['teacher'])
def get_teacher_ktp():
    teacher_id = session['user_id']
    discipline_id = request.args.get('discipline_id')
    
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    query = """
        SELECT k.ktp_id, k.discipline_id, k.group_name, k.type, k.description, k.practice_number, k.homework
        FROM ktp k
        WHERE k.teacher_id = ?
    """
    params = [teacher_id]
    
    if discipline_id:
        query += " AND k.discipline_id = ?"
        params.append(discipline_id)
    
    query += " ORDER BY k.description"
    cursor.execute(query, params)
    ktps = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(ktps)

@app.route('/api/teacher/grades/set', methods=['POST'])
@login_required
@role_required(['teacher'])
def set_grade():
    """Выставление оценки"""
    teacher_id = session['user_id']
    data = request.json
    
    student_id = data.get('student_id')
    discipline_id = data.get('discipline_id')
    ktp_id = data.get('ktp_id')
    grade = data.get('grade')
    date = data.get('date')
    
    if grade == 'н':
        grade = -1
    else:
        grade = int(grade)
    
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    cursor.execute("SELECT grade_id FROM grades WHERE student_id = ? AND discipline_id = ? AND ktp_id = ?",
                  (student_id, discipline_id, ktp_id))
    existing = cursor.fetchone()
    
    is_update = bool(existing)
    
    try:
        if existing:
            cursor.execute("""
                UPDATE grades 
                SET grade = ?, date = ? 
                WHERE student_id = ? AND discipline_id = ? AND ktp_id = ?
            """, (grade, date, student_id, discipline_id, ktp_id))
        else:
            cursor.execute("""
                INSERT INTO grades (student_id, teacher_id, discipline_id, ktp_id, grade, date)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (student_id, teacher_id, discipline_id, ktp_id, grade, date))
        
        if grade != -1:
            cursor.execute("SELECT tokens_per_attendance FROM teachers WHERE teacher_id = ?", (teacher_id,))
            result = cursor.fetchone()
            if result:
                tokens_per_attendance = result[0]
                cursor.execute("""
                    UPDATE student_teacher 
                    SET tokens = tokens + ? 
                    WHERE student_id = ? AND teacher_id = ?
                """, (tokens_per_attendance, student_id, teacher_id))
        
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'is_update': is_update})
    except Exception as e:
        logger.error(f"Error setting grade: {e}")
        conn.rollback()
        conn.close()
        return jsonify({'error': f'Database error: {str(e)}'}), 500

@app.route('/api/teacher/gradebook')
@login_required
@role_required(['teacher'])
def create_gradebook():
    """Создание ведомости с логотипом"""
    teacher_id = session['user_id']
    discipline_id = request.args.get('discipline_id')
    
    if not discipline_id:
        return jsonify({'error': 'discipline_id required'}), 400
    
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    cursor.execute("SELECT name, group_name FROM disciplines WHERE discipline_id = ?", (discipline_id,))
    discipline = cursor.fetchone()
    
    if not discipline:
        conn.close()
        return jsonify({'error': 'Discipline not found'}), 404
    
    discipline_name = discipline['name']
    group_name = discipline['group_name']
    
    cursor.execute("""
        SELECT s.student_id, u.full_name 
        FROM group_students gs
        JOIN students s ON gs.student_id = s.student_id
        JOIN users u ON s.student_id = u.user_id
        JOIN student_groups sg ON gs.group_id = sg.group_id
        WHERE sg.group_name = ? AND sg.teacher_id = ?
        ORDER BY u.full_name
    """, (group_name, teacher_id))
    students = cursor.fetchall()
    
    cursor.execute("""
        SELECT DISTINCT g.date, g.ktp_id, k.type, k.practice_number
        FROM grades g
        JOIN ktp k ON g.ktp_id = k.ktp_id
        WHERE g.discipline_id = ? AND g.teacher_id = ? AND g.grade IS NOT NULL AND g.grade != -1
        ORDER BY g.date, g.ktp_id
    """, (discipline_id, teacher_id))
    date_ktp_pairs = cursor.fetchall()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ведомость"
    
    # Логотип
    logo_path = os.path.join('static', 'logo.png')
    if os.path.exists(logo_path):
        img = Image(logo_path)
        img.width = 100
        img.height = 100
        ws.add_image(img, 'A1')
    
    ws['A3'] = f"Ведомость по дисциплине: {discipline_name}"
    ws['A3'].font = Font(bold=True)
    ws['A4'] = f"Группа: {group_name}"
    ws['A4'].font = Font(bold=True)
    
    headers = ["№", "ФИО студента"]
    ktp_counters = {"lecture": {}, "practice": {}}
    for date, ktp_id, ktp_type, practice_number in date_ktp_pairs:
        if ktp_type == "lecture":
            if date not in ktp_counters["lecture"]:
                ktp_counters["lecture"][date] = 1
            else:
                ktp_counters["lecture"][date] += 1
            header = f"{date} Лекция #{ktp_counters['lecture'][date]}"
        else:
            header = f"{date} Практика #{practice_number or 'N/A'}"
        headers.append(header)
    headers += ["Средний балл", "Посещаемость (%)"]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for row_num, student in enumerate(students, 6):
        student_id = student['student_id']
        ws.cell(row=row_num, column=1, value=row_num-5).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=2, value=student['full_name']).alignment = Alignment(horizontal='center')
        
        col_index = 3
        grades = []
        for date, ktp_id, ktp_type, practice_number in date_ktp_pairs:
            cursor.execute("""
                SELECT grade 
                FROM grades 
                WHERE student_id = ? AND discipline_id = ? AND date = ? AND ktp_id = ?
            """, (student_id, discipline_id, date, ktp_id))
            grade = cursor.fetchone()
            grade_value = str(grade['grade']) if grade and grade['grade'] != -1 else ""
            ws.cell(row=row_num, column=col_index, value=grade_value).alignment = Alignment(horizontal='center')
            if grade and grade['grade'] != -1:
                grades.append(grade['grade'])
            col_index += 1
        
        avg = sum(grades) / len(grades) if grades else 0
        ws.cell(row=row_num, column=col_index, value=round(avg, 2)).alignment = Alignment(horizontal='center')
        
        cursor.execute("""
            SELECT COUNT(*) 
            FROM grades 
            WHERE student_id = ? AND discipline_id = ? AND grade != -1
        """, (student_id, discipline_id))
        attended = cursor.fetchone()[0]
        cursor.execute("""
            SELECT COUNT(*) 
            FROM grades 
            WHERE student_id = ? AND discipline_id = ?
        """, (student_id, discipline_id))
        total = cursor.fetchone()[0]
        attendance = (attended / total * 100) if total > 0 else 0
        ws.cell(row=row_num, column=col_index + 1, value=round(attendance, 2)).alignment = Alignment(horizontal='center')
    
    for col_num in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_num)
        max_length = max(
            len(str(ws.cell(row=row, column=col_num).value or "")) for row in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    filename = f"gradebook_{discipline_name}_{group_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join('temp', filename)
    os.makedirs('temp', exist_ok=True)
    wb.save(filepath)
    conn.close()
    
    return send_file(filepath, as_attachment=True, download_name=filename)

@app.route('/api/teacher/stats')
@login_required
@role_required(['teacher'])
def get_teacher_stats():
    """Статистика для дашборда преподавателя"""
    teacher_id = session['user_id']
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    cursor.execute("SELECT COUNT(DISTINCT student_id) FROM student_teacher WHERE teacher_id = ?", (teacher_id,))
    students_count = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM disciplines WHERE teacher_id = ?", (teacher_id,))
    disciplines_count = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(DISTINCT group_name) FROM student_groups WHERE teacher_id = ?", (teacher_id,))
    groups_count = cursor.fetchone()[0]
    
    cursor.execute("""
        SELECT AVG(grade) 
        FROM grades 
        WHERE teacher_id = ? AND grade != -1 AND grade IS NOT NULL
    """, (teacher_id,))
    avg_grade = cursor.fetchone()[0] or 0
    
    cursor.execute("""
        SELECT grade, COUNT(*) as count
        FROM grades
        WHERE teacher_id = ? AND grade != -1 AND grade IS NOT NULL
        GROUP BY grade
        ORDER BY grade
    """, (teacher_id,))
    grade_distribution = {row['grade']: row['count'] for row in cursor.fetchall()}
    
    conn.close()
    
    return jsonify({
        'students_count': students_count,
        'disciplines_count': disciplines_count,
        'groups_count': groups_count,
        'avg_grade': round(avg_grade, 2),
        'grade_distribution': grade_distribution
    })

# ========== АДМИН ==========

@app.route('/admin')
@login_required
@role_required(['admin'])
def admin_dashboard():
    return render_template('admin/dashboard.html')

@app.route('/api/admin/stats')
@login_required
@role_required(['admin'])
def get_admin_stats():
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    cursor.execute("SELECT COUNT(*) FROM students")
    students_count = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM teachers")
    teachers_count = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM disciplines")
    disciplines_count = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM grades WHERE grade != -1")
    grades_count = cursor.fetchone()[0]
    
    cursor.execute("SELECT role, COUNT(*) as count FROM users GROUP BY role")
    role_distribution = {row['role']: row['count'] for row in cursor.fetchall()}
    
    conn.close()
    
    return jsonify({
        'students_count': students_count,
        'teachers_count': teachers_count,
        'disciplines_count': disciplines_count,
        'grades_count': grades_count,
        'role_distribution': role_distribution
    })

@app.route('/api/admin/teachers')
@login_required
@role_required(['admin'])
def get_admin_teachers():
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    cursor.execute("""
        SELECT t.teacher_id, u.full_name 
        FROM admin_teacher at
        JOIN teachers t ON at.teacher_id = t.teacher_id
        JOIN users u ON t.teacher_id = u.user_id
        WHERE at.admin_id = ?
    """, (session['user_id'],))
    
    teachers = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(teachers)

@app.route('/api/admin/students')
@login_required
@role_required(['admin'])
def get_admin_students():
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    cursor.execute("""
        SELECT s.student_id, u.full_name, s.group_name 
        FROM students s
        JOIN users u ON s.student_id = u.user_id
        JOIN admin_student ast ON s.student_id = ast.student_id
        WHERE ast.admin_id = ?
        ORDER BY u.full_name
    """, (session['user_id'],))
    
    students = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(students)

# ========== НОВЫЕ API ENDPOINTS ==========

@app.route('/api/teacher/grades/bulk', methods=['POST'])
@login_required
@role_required(['teacher'])
def set_bulk_grades():
    """Выставление оценок по списку"""
    teacher_id = session['user_id']
    data = request.json
    
    discipline_id = data.get('discipline_id')
    ktp_id = data.get('ktp_id')
    date = data.get('date')
    grades = data.get('grades')  # [{student_id, grade}, ...]
    
    if not all([discipline_id, ktp_id, date, grades]):
        return jsonify({'error': 'Missing required fields'}), 400
    
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    try:
        saved_count = 0
        for item in grades:
            student_id = item.get('student_id')
            grade = item.get('grade')
            
            if not student_id or grade is None:
                continue
            
            if grade == 'н' or grade == '':
                grade = -1
            else:
                grade = int(grade)
            
            cursor.execute("SELECT grade_id FROM grades WHERE student_id = ? AND discipline_id = ? AND ktp_id = ?",
                          (student_id, discipline_id, ktp_id))
            existing = cursor.fetchone()
            
            if existing:
                cursor.execute("""
                    UPDATE grades 
                    SET grade = ?, date = ? 
                    WHERE student_id = ? AND discipline_id = ? AND ktp_id = ?
                """, (grade, date, student_id, discipline_id, ktp_id))
            else:
                cursor.execute("""
                    INSERT INTO grades (student_id, teacher_id, discipline_id, ktp_id, grade, date)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (student_id, teacher_id, discipline_id, ktp_id, grade, date))
            
            if grade != -1:
                cursor.execute("SELECT tokens_per_attendance FROM teachers WHERE teacher_id = ?", (teacher_id,))
                result = cursor.fetchone()
                if result:
                    tokens_per_attendance = result[0]
                    cursor.execute("""
                        UPDATE student_teacher 
                        SET tokens = tokens + ? 
                        WHERE student_id = ? AND teacher_id = ?
                    """, (tokens_per_attendance, student_id, teacher_id))
            
            saved_count += 1
        
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'saved_count': saved_count})
    except Exception as e:
        logger.error(f"Error setting bulk grades: {e}")
        conn.rollback()
        conn.close()
        return jsonify({'error': f'Database error: {str(e)}'}), 500

@app.route('/api/teacher/grades/export', methods=['GET'])
@login_required
@role_required(['teacher'])
def export_grades_to_excel():
    """Экспорт оценок в Excel"""
    teacher_id = session['user_id']
    discipline_id = request.args.get('discipline_id')
    ktp_id = request.args.get('ktp_id')
    date = request.args.get('date')
    
    if not all([discipline_id, ktp_id, date]):
        return jsonify({'error': 'Missing parameters'}), 400
    
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    try:
        cursor.execute("SELECT name, group_name FROM disciplines WHERE discipline_id = ?", (discipline_id,))
        discipline = cursor.fetchone()
        if not discipline:
            conn.close()
            return jsonify({'error': 'Discipline not found'}), 404
        
        cursor.execute("SELECT description, type, practice_number FROM ktp WHERE ktp_id = ?", (ktp_id,))
        ktp = cursor.fetchone()
        if not ktp:
            conn.close()
            return jsonify({'error': 'KTP not found'}), 404
        
        cursor.execute("""
            SELECT s.student_id, u.full_name, g.grade
            FROM students s
            JOIN users u ON s.student_id = u.user_id
            LEFT JOIN grades g ON s.student_id = g.student_id AND g.discipline_id = ? AND g.ktp_id = ? AND g.date = ?
            JOIN student_teacher st ON s.student_id = st.student_id
            WHERE st.teacher_id = ? AND s.group_name = ?
            ORDER BY u.full_name
        """, (discipline_id, ktp_id, date, teacher_id, discipline['group_name']))
        
        students = cursor.fetchall()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Оценки"
        
        # Логотип
        logo_path = os.path.join('static', 'logo.png')
        if os.path.exists(logo_path):
            try:
                img = Image(logo_path)
                # Уменьшаем размер логотипа для нормального вписывания
                img.width = 80
                img.height = 80
                ws.add_image(img, 'A1')
            except Exception as e:
                logger.warning(f"Could not add logo: {e}")
        
        # Заголовки
        row_start = 3 if os.path.exists(logo_path) else 1
        ws.cell(row=row_start, column=1, value=f"Дисциплина: {discipline['name']}")
        ws.cell(row=row_start, column=1).font = Font(bold=True)
        ws.cell(row=row_start+1, column=1, value=f"Группа: {discipline['group_name']}")
        ws.cell(row=row_start+1, column=1).font = Font(bold=True)
        ktp_type = "Лекция" if ktp['type'] == 'lecture' else f"Практика #{ktp['practice_number']}"
        ws.cell(row=row_start+2, column=1, value=f"КТП: {ktp['description']} ({ktp_type})")
        ws.cell(row=row_start+2, column=1).font = Font(bold=True)
        ws.cell(row=row_start+3, column=1, value=f"Дата: {date}")
        ws.cell(row=row_start+3, column=1).font = Font(bold=True)
        
        headers = ["№", "ФИО студента", "Оценка"]
        header_row = row_start + 5
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for row_num, student in enumerate(students, header_row + 1):
            ws.cell(row=row_num, column=1, value=row_num - header_row).alignment = Alignment(horizontal='center')
            ws.cell(row=row_num, column=2, value=student['full_name'])
            grade_val = "н" if student['grade'] == -1 else str(student['grade']) if student['grade'] else ""
            ws.cell(row=row_num, column=3, value=grade_val).alignment = Alignment(horizontal='center')
        
        for col_num in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_num)
            max_length = max(
                len(str(ws.cell(row=row, column=col_num).value or "")) for row in range(1, ws.max_row + 1)
            )
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
        
        filename = f"grades_{discipline['name']}_{date}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join('temp', filename)
        os.makedirs('temp', exist_ok=True)
        wb.save(filepath)
        conn.close()
        
        return send_file(filepath, as_attachment=True, download_name=filename)
    except Exception as e:
        logger.error(f"Error exporting grades: {e}")
        conn.close()
        return jsonify({'error': f'Export error: {str(e)}'}), 500

@app.route('/api/teacher/disciplines/add', methods=['POST'])
@login_required
@role_required(['teacher'])
def add_discipline():
    """Добавление дисциплины"""
    teacher_id = session['user_id']
    data = request.json
    
    name = data.get('name')
    group_name = data.get('group_name')
    required_practices = data.get('required_practices', 0)
    
    if not all([name, group_name]):
        return jsonify({'error': 'Missing required fields'}), 400
    
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    try:
        cursor.execute("""
            INSERT INTO disciplines (teacher_id, name, group_name, required_practices)
            VALUES (?, ?, ?, ?)
        """, (teacher_id, name, group_name, required_practices))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'discipline_id': cursor.lastrowid})
    except Exception as e:
        logger.error(f"Error adding discipline: {e}")
        conn.rollback()
        conn.close()
        return jsonify({'error': f'Database error: {str(e)}'}), 500

@app.route('/api/teacher/disciplines/delete', methods=['POST'])
@login_required
@role_required(['teacher'])
def delete_discipline():
    """Удаление дисциплины"""
    teacher_id = session['user_id']
    data = request.json
    discipline_id = data.get('discipline_id')
    
    if not discipline_id:
        return jsonify({'error': 'discipline_id required'}), 400
    
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    try:
        cursor.execute("SELECT teacher_id FROM disciplines WHERE discipline_id = ?", (discipline_id,))
        disc = cursor.fetchone()
        if not disc or disc['teacher_id'] != teacher_id:
            conn.close()
            return jsonify({'error': 'Discipline not found or access denied'}), 404
        
        cursor.execute("DELETE FROM grades WHERE discipline_id = ?", (discipline_id,))
        cursor.execute("DELETE FROM rewards WHERE discipline_id = ?", (discipline_id,))
        cursor.execute("DELETE FROM ktp WHERE discipline_id = ?", (discipline_id,))
        cursor.execute("DELETE FROM disciplines WHERE discipline_id = ?", (discipline_id,))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Error deleting discipline: {e}")
        conn.rollback()
        conn.close()
        return jsonify({'error': f'Database error: {str(e)}'}), 500

@app.route('/api/teacher/groups/add', methods=['POST'])
@login_required
@role_required(['teacher'])
def add_group():
    """Добавление группы"""
    teacher_id = session['user_id']
    data = request.json
    group_name = data.get('group_name')
    
    if not group_name:
        return jsonify({'error': 'group_name required'}), 400
    
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    try:
        cursor.execute("INSERT OR IGNORE INTO student_groups (teacher_id, group_name) VALUES (?, ?)", 
                      (teacher_id, group_name))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Error adding group: {e}")
        conn.rollback()
        conn.close()
        return jsonify({'error': f'Database error: {str(e)}'}), 500

@app.route('/api/teacher/groups/delete', methods=['POST'])
@login_required
@role_required(['teacher'])
def delete_group():
    """Удаление группы"""
    teacher_id = session['user_id']
    data = request.json
    group_name = data.get('group_name')
    
    if not group_name:
        return jsonify({'error': 'group_name required'}), 400
    
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    try:
        cursor.execute("SELECT group_id FROM student_groups WHERE teacher_id = ? AND group_name = ?", 
                      (teacher_id, group_name))
        group = cursor.fetchone()
        
        if group:
            group_id = group[0]
            cursor.execute("DELETE FROM group_students WHERE group_id = ?", (group_id,))
            cursor.execute("DELETE FROM student_groups WHERE group_id = ?", (group_id,))
            cursor.execute("DELETE FROM disciplines WHERE teacher_id = ? AND group_name = ?", 
                          (teacher_id, group_name))
            cursor.execute("DELETE FROM ktp WHERE teacher_id = ? AND group_name = ?", 
                          (teacher_id, group_name))
            conn.commit()
        
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Error deleting group: {e}")
        conn.rollback()
        conn.close()
        return jsonify({'error': f'Database error: {str(e)}'}), 500

@app.route('/api/teacher/students/add', methods=['POST'])
@login_required
@role_required(['teacher'])
def add_student():
    """Добавление студента"""
    teacher_id = session['user_id']
    data = request.json
    
    full_name = data.get('full_name')
    group_name = data.get('group_name')
    
    if not all([full_name, group_name]):
        return jsonify({'error': 'Missing required fields'}), 400
    
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    try:
        cursor.execute("SELECT student_id FROM students WHERE full_name = ? AND group_name = ?", 
                      (full_name, group_name))
        existing = cursor.fetchone()
        
        if existing:
            student_id = existing[0]
        else:
            cursor.execute("INSERT INTO students (full_name, group_name) VALUES (?, ?)", 
                          (full_name, group_name))
            student_id = cursor.lastrowid
            cursor.execute("INSERT OR IGNORE INTO users (user_id, full_name, role, group_name) VALUES (?, ?, 'student', ?)", 
                          (student_id, full_name, group_name))
            cursor.execute("INSERT OR IGNORE INTO admin_student (admin_id, student_id) VALUES (?, ?)", 
                          (1122288946, student_id))  # ADMIN_ID
        
        cursor.execute("SELECT group_id FROM student_groups WHERE teacher_id = ? AND group_name = ?", 
                      (teacher_id, group_name))
        group = cursor.fetchone()
        
        if group:
            group_id = group[0]
            cursor.execute("INSERT OR IGNORE INTO group_students (group_id, student_id) VALUES (?, ?)", 
                          (group_id, student_id))
        
        cursor.execute("INSERT OR IGNORE INTO student_teacher (student_id, teacher_id, tokens) VALUES (?, ?, 0)", 
                      (student_id, teacher_id))
        
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'student_id': student_id})
    except Exception as e:
        logger.error(f"Error adding student: {e}")
        conn.rollback()
        conn.close()
        return jsonify({'error': f'Database error: {str(e)}'}), 500

@app.route('/api/teacher/students/delete', methods=['POST'])
@login_required
@role_required(['teacher'])
def delete_student():
    """Удаление студента из группы"""
    teacher_id = session['user_id']
    data = request.json
    student_id = data.get('student_id')
    group_name = data.get('group_name')
    
    if not all([student_id, group_name]):
        return jsonify({'error': 'Missing required fields'}), 400
    
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    try:
        cursor.execute("SELECT group_id FROM student_groups WHERE teacher_id = ? AND group_name = ?", 
                      (teacher_id, group_name))
        group = cursor.fetchone()
        
        if group:
            group_id = group[0]
            cursor.execute("DELETE FROM group_students WHERE group_id = ? AND student_id = ?", 
                          (group_id, student_id))
            conn.commit()
        
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Error deleting student: {e}")
        conn.rollback()
        conn.close()
        return jsonify({'error': f'Database error: {str(e)}'}), 500

@app.route('/api/teacher/ktp/add', methods=['POST'])
@login_required
@role_required(['teacher'])
def add_ktp():
    """Создание КТП"""
    teacher_id = session['user_id']
    data = request.json
    
    discipline_id = data.get('discipline_id')
    ktp_type = data.get('type')
    description = data.get('description')
    practice_number = data.get('practice_number')
    homework = data.get('homework')
    
    if not all([discipline_id, ktp_type, description]):
        return jsonify({'error': 'Missing required fields'}), 400
    
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    try:
        cursor.execute("SELECT group_name FROM disciplines WHERE discipline_id = ?", (discipline_id,))
        discipline = cursor.fetchone()
        if not discipline:
            conn.close()
            return jsonify({'error': 'Discipline not found'}), 404
        
        group_name = discipline['group_name']
        
        cursor.execute("""
            INSERT INTO ktp (teacher_id, discipline_id, group_name, type, description, practice_number, homework)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (teacher_id, discipline_id, group_name, ktp_type, description, 
              practice_number if ktp_type == 'practice' else None, homework or None))
        
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'ktp_id': cursor.lastrowid})
    except Exception as e:
        logger.error(f"Error adding KTP: {e}")
        conn.rollback()
        conn.close()
        return jsonify({'error': f'Database error: {str(e)}'}), 500

@app.route('/api/teacher/ktp/delete', methods=['POST'])
@login_required
@role_required(['teacher'])
def delete_ktp():
    """Удаление КТП"""
    teacher_id = session['user_id']
    data = request.json
    ktp_id = data.get('ktp_id')
    
    if not ktp_id:
        return jsonify({'error': 'ktp_id required'}), 400
    
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    try:
        cursor.execute("SELECT teacher_id FROM ktp WHERE ktp_id = ?", (ktp_id,))
        ktp = cursor.fetchone()
        if not ktp or ktp['teacher_id'] != teacher_id:
            conn.close()
            return jsonify({'error': 'KTP not found or access denied'}), 404
        
        cursor.execute("DELETE FROM grades WHERE ktp_id = ?", (ktp_id,))
        cursor.execute("DELETE FROM ktp WHERE ktp_id = ?", (ktp_id,))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Error deleting KTP: {e}")
        conn.rollback()
        conn.close()
        return jsonify({'error': f'Database error: {str(e)}'}), 500

@app.route('/api/teacher/news/send', methods=['POST'])
@login_required
@role_required(['teacher'])
def send_teacher_news():
    """Отправка новости группе"""
    teacher_id = session['user_id']
    data = request.json
    
    group_name = data.get('group_name')
    message = data.get('message')
    
    if not all([group_name, message]):
        return jsonify({'error': 'Missing required fields'}), 400
    
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    try:
        cursor.execute("SELECT full_name FROM users WHERE user_id = ?", (teacher_id,))
        teacher = cursor.fetchone()
        teacher_name = teacher['full_name'] if teacher else "Преподаватель"
        
        cursor.execute("SELECT student_id FROM students WHERE group_name = ?", (group_name,))
        students = cursor.fetchall()
        
        sent_count = 0
        for student in students:
            try:
                # Здесь нужно использовать бот для отправки, но так как бот на другом сервере,
                # можно сохранить в БД для последующей отправки или использовать API бота
                sent_count += 1
            except Exception as e:
                logger.error(f"Error sending news to student {student['student_id']}: {e}")
        
        conn.close()
        return jsonify({'success': True, 'sent_count': sent_count, 'total': len(students)})
    except Exception as e:
        logger.error(f"Error sending news: {e}")
        conn.close()
        return jsonify({'error': f'Database error: {str(e)}'}), 500

@app.route('/api/admin/students/add', methods=['POST'])
@login_required
@role_required(['admin'])
def admin_add_student():
    """Добавление студента (админ)"""
    data = request.json
    
    full_name = data.get('full_name')
    group_name = data.get('group_name')
    
    if not all([full_name, group_name]):
        return jsonify({'error': 'Missing required fields'}), 400
    
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    try:
        cursor.execute("SELECT student_id FROM students WHERE full_name = ? AND group_name = ?", 
                      (full_name, group_name))
        existing = cursor.fetchone()
        
        if existing:
            student_id = existing[0]
        else:
            cursor.execute("INSERT INTO students (full_name, group_name) VALUES (?, ?)", 
                          (full_name, group_name))
            student_id = cursor.lastrowid
            cursor.execute("INSERT OR IGNORE INTO users (user_id, full_name, role, group_name) VALUES (?, ?, 'student', ?)", 
                          (student_id, full_name, group_name))
        
        cursor.execute("INSERT OR IGNORE INTO admin_student (admin_id, student_id) VALUES (?, ?)", 
                      (session['user_id'], student_id))
        
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'student_id': student_id})
    except Exception as e:
        logger.error(f"Error adding student: {e}")
        conn.rollback()
        conn.close()
        return jsonify({'error': f'Database error: {str(e)}'}), 500

@app.route('/api/admin/students/delete', methods=['POST'])
@login_required
@role_required(['admin'])
def admin_delete_student():
    """Удаление студента (админ)"""
    data = request.json
    student_id = data.get('student_id')
    
    if not student_id:
        return jsonify({'error': 'student_id required'}), 400
    
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    try:
        cursor.execute("DELETE FROM admin_student WHERE student_id = ?", (student_id,))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Error deleting student: {e}")
        conn.rollback()
        conn.close()
        return jsonify({'error': f'Database error: {str(e)}'}), 500

@app.route('/api/admin/news/send', methods=['POST'])
@login_required
@role_required(['admin'])
def send_admin_news():
    """Отправка новости (админ)"""
    data = request.json
    
    recipient = data.get('recipient')  # 'students', 'teachers', 'all'
    message = data.get('message')
    
    if not all([recipient, message]):
        return jsonify({'error': 'Missing required fields'}), 400
    
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    try:
        if recipient == 'all':
            cursor.execute("SELECT user_id FROM users WHERE role IN ('student', 'teacher')")
        elif recipient == 'students':
            cursor.execute("SELECT user_id FROM users WHERE role = 'student'")
        else:
            cursor.execute("SELECT user_id FROM users WHERE role = 'teacher'")
        
        users = cursor.fetchall()
        sent_count = len(users)  # В реальности нужно отправлять через бот
        
        conn.close()
        return jsonify({'success': True, 'sent_count': sent_count})
    except Exception as e:
        logger.error(f"Error sending news: {e}")
        conn.close()
        return jsonify({'error': f'Database error: {str(e)}'}), 500

if __name__ == '__main__':
    # Для разработки используйте debug=True
    # Для продакшена установите debug=False
    import os
    debug_mode = os.getenv('FLASK_ENV') != 'production'
    app.run(debug=debug_mode, host='0.0.0.0', port=int(os.getenv('PORT', 5000)))