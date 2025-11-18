import os
import sqlite3
import pandas as pd
from datetime import datetime, timedelta
from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import (
    Message,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    ReplyKeyboardMarkup,
    KeyboardButton,
    ReplyKeyboardRemove,
    FSInputFile
)
from aiogram.utils.keyboard import InlineKeyboardBuilder
import logging
import uuid
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

async def notify_student(student_id: int, discipline_name: str, grade: int, date: str, ktp_description: str, teacher_name: str, homework: str = None, is_update: bool = False):
    grade_str = "н" if grade == -1 else str(grade)
    action = "изменена" if is_update else "выставлена"
    message = f"Вам {action} оценка по дисциплине '{discipline_name}' за {date} (КТП: {ktp_description}): {grade_str} (Преподаватель: {teacher_name})"
    if homework:
        message += f"\nДомашнее задание: {homework}"
    try:
        await bot.send_message(chat_id=student_id, text=message)
    except Exception as e:
        logger.error(f"Ошибка отправки уведомления студенту {student_id}: {e}")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

BOT_TOKEN = "8442631201:AAGkUM0m7DE1wk4G1Vt0tDjIS-RbL4bZaQo"
ADMIN_ID = 1122288946

bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()

conn = sqlite3.connect("students.db", check_same_thread=False)
cursor = conn.cursor()

def init_db():
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS first_practice_completions (
        discipline_id INTEGER,
        student_id INTEGER,
        tokens INTEGER,
        awarded_date TEXT,
        PRIMARY KEY (discipline_id),
        FOREIGN KEY (discipline_id) REFERENCES disciplines(discipline_id),
        FOREIGN KEY (student_id) REFERENCES students(student_id)
    )
    """)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS reserved_tokens (
        student_id INTEGER,
        teacher_id INTEGER,
        discipline_id INTEGER,
        tokens INTEGER DEFAULT 0,
        notification_message TEXT NOT NULL,
        PRIMARY KEY (student_id, teacher_id, discipline_id),
        FOREIGN KEY (student_id) REFERENCES students(student_id),
        FOREIGN KEY (teacher_id) REFERENCES teachers(teacher_id),
        FOREIGN KEY (discipline_id) REFERENCES disciplines(discipline_id)
    )
    """)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS users (
        user_id INTEGER PRIMARY KEY,
        username TEXT,
        full_name TEXT,
        role TEXT CHECK(role IN ('student', 'teacher', 'admin')),
        group_name TEXT,
        token TEXT UNIQUE
    )
    """)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS teachers (
        teacher_id INTEGER PRIMARY KEY,
        full_name TEXT,
        token TEXT UNIQUE,
        tokens_per_attendance INTEGER DEFAULT 1,
        FOREIGN KEY (teacher_id) REFERENCES users(user_id)
    )
    """)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS students (
        student_id INTEGER PRIMARY KEY,
        full_name TEXT,
        group_name TEXT,
        FOREIGN KEY (student_id) REFERENCES users(user_id)
    )
    """)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS student_teacher (
        student_id INTEGER,
        teacher_id INTEGER,
        tokens INTEGER DEFAULT 0,
        PRIMARY KEY (student_id, teacher_id),
        FOREIGN KEY (student_id) REFERENCES students(student_id),
        FOREIGN KEY (teacher_id) REFERENCES teachers(teacher_id)
    )
    """)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS admin_teacher (
        admin_id INTEGER,
        teacher_id INTEGER,
        PRIMARY KEY (admin_id, teacher_id),
        FOREIGN KEY (admin_id) REFERENCES users(user_id),
        FOREIGN KEY (teacher_id) REFERENCES teachers(teacher_id)
    )
    """)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS admin_student (
        admin_id INTEGER,
        student_id INTEGER,
        PRIMARY KEY (admin_id, student_id),
        FOREIGN KEY (admin_id) REFERENCES users(user_id),
        FOREIGN KEY (student_id) REFERENCES students(student_id)
    )
    """)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS disciplines (
        discipline_id INTEGER PRIMARY KEY AUTOINCREMENT,
        teacher_id INTEGER,
        name TEXT,
        group_name TEXT,
        FOREIGN KEY (teacher_id) REFERENCES teachers(teacher_id)
    )
    """)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS ktp (
        ktp_id INTEGER PRIMARY KEY AUTOINCREMENT,
        teacher_id INTEGER,
        discipline_id INTEGER,
        group_name TEXT,
        type TEXT CHECK(type IN ('lecture', 'practice')),
        description TEXT,
        practice_number INTEGER,
        homework TEXT,
        FOREIGN KEY (teacher_id) REFERENCES teachers(teacher_id),
        FOREIGN KEY (discipline_id) REFERENCES disciplines(discipline_id)
    )
    """)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS grades (
        grade_id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id INTEGER,
        teacher_id INTEGER,
        discipline_id INTEGER,
        ktp_id INTEGER,
        grade INTEGER CHECK(grade IN (0, 1, 2, 3, 4, 5, -1)),
        date TEXT,
        FOREIGN KEY (student_id) REFERENCES students(student_id),
        FOREIGN KEY (teacher_id) REFERENCES teachers(teacher_id),
        FOREIGN KEY (discipline_id) REFERENCES disciplines(discipline_id),
        FOREIGN KEY (ktp_id) REFERENCES ktp(ktp_id)
    )
    """)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS student_groups (
        group_id INTEGER PRIMARY KEY AUTOINCREMENT,
        teacher_id INTEGER,
        group_name TEXT,
        FOREIGN KEY (teacher_id) REFERENCES teachers(teacher_id)
    )
    """)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS group_students (
        group_id INTEGER,
        student_id INTEGER,
        PRIMARY KEY (group_id, student_id),
        FOREIGN KEY (group_id) REFERENCES student_groups(group_id),
        FOREIGN KEY (student_id) REFERENCES students(student_id)
    )
    """)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS rewards (
        reward_id INTEGER PRIMARY KEY AUTOINCREMENT,
        teacher_id INTEGER,
        discipline_id INTEGER,
        name TEXT,
        description TEXT,
        price INTEGER,
        is_enabled INTEGER DEFAULT 1,
        FOREIGN KEY (teacher_id) REFERENCES teachers(teacher_id),
        FOREIGN KEY (discipline_id) REFERENCES disciplines(discipline_id)
    )
    """)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS purchased_rewards (
        purchase_id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id INTEGER,
        reward_id INTEGER,
        teacher_id INTEGER,
        purchase_date TEXT,
        FOREIGN KEY (student_id) REFERENCES students(student_id),
        FOREIGN KEY (reward_id) REFERENCES rewards(reward_id),
        FOREIGN KEY (teacher_id) REFERENCES teachers(teacher_id)
    )
    """)
    
    try:
        cursor.execute("ALTER TABLE disciplines ADD COLUMN required_practices INTEGER")
        logger.info("Столбец 'required_practices' успешно добавлен в таблицу disciplines.")
    except sqlite3.OperationalError as e:
        if "duplicate column name" in str(e):
            logger.info("Столбец 'required_practices' уже существует в таблице disciplines.")
        else:
            raise

    # Миграция: добавление столбца homework
    try:
        cursor.execute("ALTER TABLE ktp ADD COLUMN homework TEXT")
        logger.info("Столбец 'homework' успешно добавлен в таблицу ktp.")
    except sqlite3.OperationalError as e:
        if "duplicate column name" in str(e):
            logger.info("Столбец 'homework' уже существует в таблице ktp.")
        else:
            logger.error(f"Ошибка при добавлении столбца 'homework': {e}")
            raise

    # Миграция: добавление столбца type
    try:
        cursor.execute("ALTER TABLE ktp ADD COLUMN type TEXT CHECK(type IN ('lecture', 'practice'))")
        logger.info("Столбец 'type' успешно добавлен в таблицу ktp.")
    except sqlite3.OperationalError as e:
        if "duplicate column name" in str(e):
            logger.info("Столбец 'type' уже существует в таблице ktp.")
        else:
            logger.error(f"Ошибка при добавлении столбца 'type': {e}")
            raise

    # Миграция: добавление столбца practice_number
    try:
        cursor.execute("ALTER TABLE ktp ADD COLUMN practice_number INTEGER")
        logger.info("Столбец 'practice_number' успешно добавлен в таблицу ktp.")
    except sqlite3.OperationalError as e:
        if "duplicate column name" in str(e):
            logger.info("Столбец 'practice_number' уже существует в таблице ktp.")
        else:
            logger.error(f"Ошибка при добавлении столбца 'practice_number': {e}")
            raise
    
    conn.commit()
    init_db()

class Form(StatesGroup):
    waiting_for_full_name = State()
    waiting_for_group = State()

class TeacherForm(StatesGroup):
    waiting_for_student_list = State()
    waiting_for_group_name = State()
    waiting_for_discipline_name = State()
    waiting_for_student_name = State()
    waiting_for_student_group = State()
    waiting_for_reward_name = State()
    waiting_for_reward_description = State()
    waiting_for_reward_price = State()
    waiting_for_edit_reward_price = State()
    waiting_for_ktp_discipline = State()
    waiting_for_ktp_type = State()
    waiting_for_lecture_topic = State()
    waiting_for_practice_number = State()
    waiting_for_practice_topic = State()
    waiting_for_tokens_per_attendance = State()
    waiting_for_view_ktp_discipline = State()
    waiting_for_lecture_homework = State()  # Новое состояние для домашнего задания лекции
    waiting_for_practice_homework = State()
    waiting_for_required_practices = State()
    waiting_for_required_practices = State()
    waiting_for_group_news = State()
    waiting_for_news_message = State()

class GradeForm(StatesGroup):
    selecting_grading_type = State()
    selecting_discipline = State()
    selecting_ktp_type = State()
    selecting_ktp = State()
    selecting_date = State()
    selecting_student = State()
    entering_grade = State()
    selecting_student_single = State()
    selecting_date_edit = State()
    selecting_ktp_edit = State()
    selecting_discipline_edit = State()
    selecting_student_edit = State()
    entering_grade_edit = State()
    
class AdminForm(StatesGroup):
    waiting_for_recipient = State()
    waiting_for_news_message = State()

# Функция проверки завершения всех практик
def check_all_practices_completed(student_id: int, discipline_id: int, teacher_id: int) -> bool:
    cursor.execute("SELECT required_practices FROM disciplines WHERE discipline_id = ?", (discipline_id,))
    required_practices = cursor.fetchone()[0]
    
    if not required_practices or required_practices < 1:
        return False

    cursor.execute("""
    SELECT COUNT(DISTINCT k.practice_number)
    FROM grades g
    JOIN ktp k ON g.ktp_id = k.ktp_id
    WHERE g.student_id = ? AND g.discipline_id = ? AND k.type = 'practice' 
    AND g.grade >= 1 AND g.grade != -1
    """, (student_id, discipline_id))
    
    completed_practices = cursor.fetchone()[0]
    return completed_practices >= required_practices

async def award_practice_completion(student_id: int, discipline_id: int, teacher_id: int):
    # Получаем название дисциплины
    cursor.execute("SELECT name FROM disciplines WHERE discipline_id = ?", (discipline_id,))
    discipline_name_result = cursor.fetchone()
    if not discipline_name_result:
        logger.error(f"Дисциплина с ID {discipline_id} не найдена")
        return
    discipline_name = discipline_name_result[0]

    # Получаем ФИО преподавателя
    cursor.execute("SELECT full_name FROM users WHERE user_id = ?", (teacher_id,))
    teacher_name_result = cursor.fetchone()
    if not teacher_name_result:
        logger.error(f"Преподаватель с ID {teacher_id} не найден")
        return
    teacher_name = teacher_name_result[0]

    # Получаем ФИО студента
    cursor.execute("SELECT full_name FROM users WHERE user_id = ? AND role = 'student'", (student_id,))
    student_name_result = cursor.fetchone()
    if not student_name_result:
        logger.error(f"Студент с ID {student_id} не зарегистрирован")
        student_name = f"Студент ID {student_id}"  # Запасное имя для уведомления
    else:
        student_name = student_name_result[0]

    # Проверяем, не был ли уже кто-то отмечен как первый сдавший
    cursor.execute("SELECT student_id FROM first_practice_completions WHERE discipline_id = ?", (discipline_id,))
    first_completion = cursor.fetchone()

    if first_completion:
        logger.info(f"Первый сдавший для дисциплины {discipline_id} уже есть: {first_completion[0]}")
        return  # Если кто-то уже сдал первым, выходим

    # Получаем дату последней сданной практики и соответствующую дату КТП
    cursor.execute("""
        SELECT g.date, k.practice_number
        FROM grades g
        JOIN ktp k ON g.ktp_id = k.ktp_id
        WHERE g.student_id = ? AND g.discipline_id = ? AND k.type = 'practice'
        AND g.grade >= 1 AND g.grade != -1
        ORDER BY g.date DESC
        LIMIT 1
    """, (student_id, discipline_id))
    last_grade = cursor.fetchone()

    if not last_grade:
        logger.info(f"У студента {student_id} нет сданных практик для дисциплины {discipline_id}")
        return

    grade_date, practice_number = last_grade

    # Получаем дату практики из КТП
    cursor.execute("""
        SELECT date
        FROM grades
        WHERE discipline_id = ? AND ktp_id IN (
            SELECT ktp_id FROM ktp WHERE discipline_id = ? AND practice_number = ? AND type = 'practice'
        )
        LIMIT 1
    """, (discipline_id, discipline_id, practice_number))
    practice_date_result = cursor.fetchone()

    if not practice_date_result:
        logger.error(f"Дата практики для дисциплины {discipline_id} и номера {practice_number} не найдена")
        return

    practice_date = practice_date_result[0]

    # Проверяем, совпадает ли дата сдачи с датой практики
    try:
        grade_date_obj = datetime.strptime(grade_date, "%d-%m-%Y")
        practice_date_obj = datetime.strptime(practice_date, "%d-%m-%Y")
        if grade_date_obj.date() != practice_date_obj.date():
            logger.info(f"Студент {student_id} сдал практику не в день практики: {grade_date} != {practice_date}")
            return
    except ValueError as e:
        logger.error(f"Ошибка формата даты: {e}")
        return

    # Проверяем, все ли практики сданы
    if not check_all_practices_completed(student_id, discipline_id, teacher_id):
        logger.info(f"Студент {student_id} не сдал все практики для дисциплины {discipline_id}")
        return

    # Если все проверки пройдены, начисляем жетоны
    tokens = 15
    message = f"Поздравляем! Вы сдали все практики по дисциплине '{discipline_name}' первыми и вовремя. Вам начислено {tokens} жетончиков."

    # Регистрируем студента как первого сдавшего
    cursor.execute("""
        INSERT OR IGNORE INTO first_practice_completions (discipline_id, student_id, tokens, awarded_date)
        VALUES (?, ?, ?, ?)
    """, (discipline_id, student_id, tokens, datetime.now().strftime("%d-%m-%Y")))

    # Проверяем, зарегистрирован ли студент
    cursor.execute("SELECT user_id FROM users WHERE user_id = ? AND role = 'student'", (student_id,))
    if cursor.fetchone():
        cursor.execute("""
            UPDATE student_teacher 
            SET tokens = tokens + ? 
            WHERE student_id = ? AND teacher_id = ?
        """, (tokens, student_id, teacher_id))
        try:
            await bot.send_message(student_id, message)
            await bot.send_message(teacher_id, f"Студент {student_name} сдал все практики по '{discipline_name}' первым и вовремя, получил {tokens} жетончиков.")
        except Exception as e:
            logger.error(f"Ошибка отправки уведомления студенту {student_id} или преподавателю {teacher_id}: {e}")
    else:
        cursor.execute("""
            INSERT OR REPLACE INTO reserved_tokens (student_id, teacher_id, discipline_id, tokens, notification_message)
            VALUES (?, ?, ?, ?, ?)
        """, (student_id, teacher_id, discipline_id, tokens, message))
    
    conn.commit()

def generate_token():
    return str(uuid.uuid4())[:8]

def is_admin(user_id: int) -> bool:
    cursor.execute("SELECT role FROM users WHERE user_id = ?", (user_id,))
    result = cursor.fetchone()
    return result and result[0] == "admin"

def is_teacher(user_id: int) -> bool:
    cursor.execute("SELECT role FROM users WHERE user_id = ?", (user_id,))
    result = cursor.fetchone()
    return result and result[0] == "teacher"

def is_student(user_id: int) -> bool:
    cursor.execute("SELECT role FROM users WHERE user_id = ?", (user_id,))
    result = cursor.fetchone()
    return result and result[0] == "student"

def get_user_info(user_id: int):
    cursor.execute("SELECT * FROM users WHERE user_id = ?", (user_id,))
    return cursor.fetchone()

def get_student_teachers(student_id: int):
    cursor.execute("""
    SELECT t.teacher_id, u.full_name, st.tokens 
    FROM student_teacher st
    JOIN teachers t ON st.teacher_id = t.teacher_id
    JOIN users u ON t.teacher_id = u.user_id
    WHERE st.student_id = ?
    """, (student_id,))
    return cursor.fetchall()

def get_admin_teachers(admin_id: int):
    cursor.execute("""
    SELECT t.teacher_id, u.full_name 
    FROM admin_teacher at
    JOIN teachers t ON at.teacher_id = t.teacher_id
    JOIN users u ON t.teacher_id = u.user_id
    WHERE at.admin_id = ?
    """, (admin_id,))
    return cursor.fetchall()

def get_admin_students(admin_id: int):
    cursor.execute("""
    SELECT s.student_id, u.full_name, s.group_name 
    FROM students s
    JOIN users u ON s.student_id = u.user_id
    JOIN admin_student ast ON s.student_id = ast.student_id
    WHERE ast.admin_id = ?
    ORDER BY u.full_name
    """, (admin_id,))
    return cursor.fetchall()

def get_teacher_students(teacher_id: int, group_name: str = None):
    query = """
    SELECT DISTINCT s.student_id, u.full_name, s.group_name 
    FROM students s
    JOIN users u ON s.student_id = u.user_id
    JOIN student_teacher st ON s.student_id = st.student_id
    WHERE st.teacher_id = ?
    """
    params = [teacher_id]
    if group_name:
        query += " AND s.group_name = ?"
        params.append(group_name)
    query += " ORDER BY u.full_name"
    cursor.execute(query, params)
    return cursor.fetchall()

def get_teacher_disciplines(teacher_id: int):
    cursor.execute("SELECT discipline_id, name, group_name FROM disciplines WHERE teacher_id = ? ORDER BY name", (teacher_id,))
    return cursor.fetchall()

def get_teacher_groups(teacher_id: int):
    cursor.execute("SELECT DISTINCT group_name FROM student_groups WHERE teacher_id = ? ORDER BY group_name", (teacher_id,))
    return [row[0] for row in cursor.fetchall()]

def get_student_grades(student_id: int):
    cursor.execute("""
    SELECT g.grade, g.date, d.name, u.full_name, COALESCE(k.description, 'Нет описания КТП') as description, COALESCE(k.homework, 'Нет домашнего задания') as homework
    FROM grades g
    JOIN disciplines d ON g.discipline_id = d.discipline_id
    JOIN users u ON g.teacher_id = u.user_id
    LEFT JOIN ktp k ON g.ktp_id = k.ktp_id
    WHERE g.student_id = ?
    ORDER BY g.date DESC
    """, (student_id,))
    grades = cursor.fetchall()
    logger.info(f"All grades for student {student_id}: {grades}")
    return grades

def get_student_info(student_id: int):
    cursor.execute("""
    SELECT s.student_id, u.full_name, s.group_name 
    FROM students s
    JOIN users u ON s.student_id = u.user_id
    WHERE s.student_id = ?
    """, (student_id,))
    return cursor.fetchone()

def get_teacher_info(teacher_id: int):
    cursor.execute("""
    SELECT t.teacher_id, u.full_name 
    FROM teachers t
    JOIN users u ON t.teacher_id = u.user_id
    WHERE t.teacher_id = ?
    """, (teacher_id,))
    return cursor.fetchone()

def is_student_linked(student_id: int, teacher_id: int):
    cursor.execute("SELECT 1 FROM student_teacher WHERE student_id = ? AND teacher_id = ?", (student_id, teacher_id))
    return cursor.fetchone() is not None

def is_teacher_linked(admin_id: int, teacher_id: int):
    cursor.execute("SELECT 1 FROM admin_teacher WHERE admin_id = ? AND teacher_id = ?", (admin_id, teacher_id))
    return cursor.fetchone() is not None

def is_student_linked_to_admin(admin_id: int, student_id: int):
    cursor.execute("SELECT 1 FROM admin_student WHERE admin_id = ? AND student_id = ?", (admin_id, student_id))
    return cursor.fetchone() is not None

def get_group_students(group_name: str, teacher_id: int):
    cursor.execute("""
    SELECT s.student_id, u.full_name 
    FROM group_students gs
    JOIN students s ON gs.student_id = s.student_id
    JOIN users u ON s.student_id = u.user_id
    JOIN student_groups sg ON gs.group_id = sg.group_id
    WHERE sg.group_name = ? AND sg.teacher_id = ?
    ORDER BY u.full_name
    """, (group_name, teacher_id))
    return cursor.fetchall()

def get_teacher_rewards(teacher_id: int):
    cursor.execute("""
    SELECT r.reward_id, r.name, r.description, r.price, r.is_enabled, d.name
    FROM rewards r
    JOIN disciplines d ON r.discipline_id = d.discipline_id
    WHERE r.teacher_id = ?
    ORDER BY d.name, r.name
    """, (teacher_id,))
    return cursor.fetchall()

def get_student_rewards(student_id: int):
    cursor.execute("""
    SELECT r.reward_id, r.name, r.description, r.price, u.full_name, d.name, r.teacher_id
    FROM rewards r
    JOIN disciplines d ON r.discipline_id = d.discipline_id
    JOIN student_teacher st ON r.teacher_id = st.teacher_id
    JOIN users u ON r.teacher_id = u.user_id
    WHERE st.student_id = ? AND r.is_enabled = 1
    ORDER BY d.name, r.name
    """, (student_id,))
    return cursor.fetchall()

def get_ktp_by_discipline_and_type(teacher_id: int, discipline_id: int, ktp_type: str):
    query = """
    SELECT k.ktp_id, k.group_name, k.type, k.description, k.practice_number
    FROM ktp k
    WHERE k.teacher_id = ? AND k.discipline_id = ? AND k.type = ?
    """
    params = [teacher_id, discipline_id, ktp_type]
    query += " ORDER BY k.description"
    cursor.execute(query, params)
    return cursor.fetchall()

def get_student_token_balance(student_id: int, teacher_id: int):
    cursor.execute("SELECT tokens FROM student_teacher WHERE student_id = ? AND teacher_id = ?", (student_id, teacher_id))
    result = cursor.fetchone()
    return result[0] if result else 0

def get_student_attendance_percentage(student_id: int, discipline_id: int):
    cursor.execute("""
    SELECT COUNT(*) 
    FROM grades 
    WHERE student_id = ? AND discipline_id = ? AND grade != -1
    """, (student_id, discipline_id))
    attended = cursor.fetchone()[0]
    
    cursor.execute("""
    SELECT COUNT(*) 
    FROM grades 
    WHERE student_id = ? AND discipline_id = ?
    """, (student_id, discipline_id))
    total = cursor.fetchone()[0]
    
    return (attended / total * 100) if total > 0 else 0

def create_grades_excel(student_id: int):
    grades = get_student_grades(student_id)
    df = pd.DataFrame(grades, columns=["Оценка", "Дата", "Дисциплина", "Преподаватель", "Описание КТП", "Домашнее задание"])
    df["Оценка"] = df["Оценка"].replace(-1, "н")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Оценки"
    
    headers = ["Оценка", "Дата", "Дисциплина", "Преподаватель", "Описание КТП", "Домашнее задание"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    for row_num, row_data in enumerate(grades, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value="н" if value == -1 else value)
            cell.alignment = Alignment(horizontal='center')
    
    for col_num in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_num)
        max_length = max(
            len(str(ws.cell(row=row, column=col_num).value or "")) for row in range(1, ws.max_row + 1)
        )
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width
    
    filename = f"grades_{student_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    return filename

def create_teachers_template():
    try:
        # Создаем DataFrame с заголовком
        df = pd.DataFrame(columns=["ФИО преподавателя"])
        filename = f"teachers_template_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Сохраняем DataFrame в Excel
        df.to_excel(filename, index=False, engine='openpyxl')
        
        # Проверяем, создан ли файл
        if not os.path.exists(filename):
            raise ValueError("Файл шаблона не был создан.")
        
        # Открываем файл для настройки
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        ws.title = "Шаблон преподавателей"
        
        # Устанавливаем ширину столбца
        ws.column_dimensions['A'].width = 30
        
        # Проверяем заголовок
        if ws['A1'].value != "ФИО преподавателя":
            logger.error("Заголовок 'ФИО преподавателя' не установлен корректно")
            raise ValueError("Некорректный заголовок в шаблоне преподавателей")
        
        # Сохраняем файл
        wb.save(filename)
        logger.info(f"Шаблон преподавателей успешно создан: {filename}")
        return filename
    
    except Exception as e:
        logger.error(f"Ошибка при создании шаблона преподавателей: {e}")
        if os.path.exists(filename):
            os.remove(filename)
        raise ValueError(f"Не удалось создать шаблон преподавателей: {str(e)}")

def create_students_template():
    try:
        # Создаем DataFrame с заголовками
        df = pd.DataFrame(columns=["ФИО студента", "Группа"])
        filename = f"students_template_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Сохраняем DataFrame в Excel
        df.to_excel(filename, index=False, engine='openpyxl')
        
        # Проверяем, создан ли файл
        if not os.path.exists(filename):
            raise ValueError("Файл шаблона не был создан.")
        
        # Открываем файл для настройки
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        ws.title = "Шаблон студентов"
        
        # Устанавливаем ширину столбцов
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        
        # Проверяем заголовки
        if ws['A1'].value != "ФИО студента" or ws['B1'].value != "Группа":
            logger.error("Заголовки 'ФИО студента' или 'Группа' не установлены корректно")
            raise ValueError("Некорректные заголовки в шаблоне студентов")
        
        # Сохраняем файл
        wb.save(filename)
        logger.info(f"Шаблон студентов успешно создан: {filename}")
        return filename
    
    except Exception as e:
        logger.error(f"Ошибка при создании шаблона студентов: {e}")
        if os.path.exists(filename):
            os.remove(filename)
        raise ValueError(f"Не удалось создать шаблон студентов: {str(e)}")

def create_gradebook(teacher_id: int, discipline_id: int):
    cursor.execute("SELECT name, group_name FROM disciplines WHERE discipline_id = ?", (discipline_id,))
    discipline_name, group_name = cursor.fetchone()
    
    students = get_group_students(group_name, teacher_id)
    
    # Получаем уникальные комбинации date и ktp_id, где есть оценки
    cursor.execute("""
    SELECT DISTINCT g.date, g.ktp_id, k.type, k.practice_number
    FROM grades g
    JOIN ktp k ON g.ktp_id = k.ktp_id
    WHERE g.discipline_id = ? AND g.teacher_id = ? AND g.grade IS NOT NULL AND g.grade != -1
    ORDER BY g.date, g.ktp_id
    """, (discipline_id, teacher_id))
    date_ktp_pairs = cursor.fetchall()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ведомость"
    
    ws['A1'] = f"Ведомость по дисциплине: {discipline_name}"
    ws['A1'].font = Font(bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A2'] = f"Группа: {group_name}"
    ws['A2'].font = Font(bold=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Формируем заголовки: номер, ФИО, даты с номерами КТП, средний балл, посещаемость
    headers = ["№", "ФИО студента"]
    ktp_counters = {"lecture": {}, "practice": {}}  # Счетчики для лекций и практик по датам
    for date, ktp_id, ktp_type, practice_number in date_ktp_pairs:
        if ktp_type == "lecture":
            if date not in ktp_counters["lecture"]:
                ktp_counters["lecture"][date] = 1
            else:
                ktp_counters["lecture"][date] += 1
            header = f"{date} Лекция #{ktp_counters['lecture'][date]}"
        else:  # practice
            header = f"{date} Практика #{practice_number or 'N/A'}"
        headers.append(header)
    headers += ["Средний балл", "Посещаемость (%)"]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    student_averages = {}
    for row_num, (student_id, full_name) in enumerate(students, 5):
        ws.cell(row=row_num, column=1, value=row_num-4).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=2, value=full_name).alignment = Alignment(horizontal='center')
        
        col_index = 3
        grades = []
        for date, ktp_id, ktp_type, practice_number in date_ktp_pairs:
            cursor.execute("""
            SELECT grade 
            FROM grades 
            WHERE student_id = ? AND discipline_id = ? AND date = ? AND ktp_id = ?
            """, (student_id, discipline_id, date, ktp_id))
            grade = cursor.fetchone()
            grade_value = str(grade[0]) if grade and grade[0] != -1 else ""
            ws.cell(row=row_num, column=col_index, value=grade_value).alignment = Alignment(horizontal='center')
            if grade and grade[0] != -1:
                grades.append(grade[0])
            col_index += 1
        
        avg = sum(grades) / len(grades) if grades else 0
        student_averages[student_id] = round(avg, 2)
        ws.cell(row=row_num, column=col_index, value=student_averages[student_id]).alignment = Alignment(horizontal='center')
        attendance = round(get_student_attendance_percentage(student_id, discipline_id), 2)
        ws.cell(row=row_num, column=col_index + 1, value=attendance).alignment = Alignment(horizontal='center')
    
    for col_num in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_num)
        max_length = max(
            len(str(ws.cell(row=row, column=col_num).value or "")) for row in range(1, ws.max_row + 1)
        )
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width
    
    filename = f"gradebook_{discipline_name}_{group_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    return filename

@dp.message(Command("start"))
async def cmd_start(message: Message, state: FSMContext):
    user_id = message.from_user.id
    if user_id == ADMIN_ID:
        cursor.execute("INSERT OR IGNORE INTO users (user_id, role) VALUES (?, 'admin')", (user_id,))
        conn.commit()
    
    user_info = get_user_info(user_id)
    
    if user_info:
        role = user_info[3]
        if role == "admin":
            await show_admin_menu(message)
        elif role == "teacher":
            await show_teacher_menu(message)
        elif role == "student":
            await show_student_menu(message)
    else:
        keyboard = ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="Я студент")],
                [KeyboardButton(text="Я преподаватель")]
            ],
            resize_keyboard=True
        )
        await message.answer("Добро пожаловать! Выберите вашу роль:", reply_markup=keyboard)

@dp.message(F.text.in_(["Я студент", "Я преподаватель"]))
async def role_selected(message: Message, state: FSMContext):
    if message.text == "Я студент":
        await state.set_state(Form.waiting_for_full_name)
        await message.answer("Введите ваше ФИО:", reply_markup=ReplyKeyboardRemove())
    elif message.text == "Я преподаватель":
        await message.answer("Введите токен, выданный администратором:", reply_markup=ReplyKeyboardRemove())

@dp.message(Form.waiting_for_full_name)
async def process_student_name(message: Message, state: FSMContext):
    await state.update_data(full_name=message.text.strip())
    await state.set_state(Form.waiting_for_group)
    await message.answer("Введите вашу группу:")

# Обновление функции регистрации студента для начисления зарезервированных жетонов
@dp.message(Form.waiting_for_group)
async def process_student_group(message: Message, state: FSMContext):
    user_id = message.from_user.id
    data = await state.get_data()
    full_name = data.get("full_name")
    group_name = message.text.strip()
    
    cursor.execute("SELECT student_id FROM students WHERE full_name = ? AND group_name = ?", (full_name, group_name))
    student = cursor.fetchone()
    
    if student:
        old_student_id = student[0]
        cursor.execute("UPDATE students SET student_id = ? WHERE student_id = ?", (user_id, old_student_id))
        cursor.execute("UPDATE users SET user_id = ?, full_name = ?, group_name = ?, role = 'student' WHERE user_id = ?",
                      (user_id, full_name, group_name, old_student_id))
        cursor.execute("UPDATE student_teacher SET student_id = ? WHERE student_id = ?", (user_id, old_student_id))
        cursor.execute("UPDATE admin_student SET student_id = ? WHERE student_id = ?", (user_id, old_student_id))
        cursor.execute("UPDATE group_students SET student_id = ? WHERE student_id = ?", (user_id, old_student_id))
        cursor.execute("UPDATE grades SET student_id = ? WHERE student_id = ?", (user_id, old_student_id))
        cursor.execute("UPDATE purchased_rewards SET student_id = ? WHERE student_id = ?", (user_id, old_student_id))
        cursor.execute("UPDATE reserved_tokens SET student_id = ? WHERE student_id = ?", (user_id, old_student_id))
        student_id = user_id
    else:
        cursor.execute("SELECT student_id FROM students WHERE full_name = ?", (full_name,))
        existing_student = cursor.fetchone()
        if existing_student:
            student_id = existing_student[0]
            cursor.execute("UPDATE students SET group_name = ? WHERE student_id = ?", (group_name, student_id))
            cursor.execute("UPDATE users SET user_id = ?, full_name = ?, group_name = ?, role = 'student' WHERE user_id = ?",
                          (user_id, full_name, group_name, student_id))
        else:
            cursor.execute("INSERT INTO users (user_id, username, full_name, role, group_name) VALUES (?, ?, ?, 'student', ?)", 
                          (user_id, message.from_user.username, full_name, group_name))
            cursor.execute("INSERT INTO students (student_id, full_name, group_name) VALUES (?, ?, ?)", 
                          (user_id, full_name, group_name))
            student_id = user_id
    
    cursor.execute("INSERT OR IGNORE INTO admin_student (admin_id, student_id) VALUES (?, ?)", 
                  (ADMIN_ID, student_id))
    
    cursor.execute("""
    SELECT DISTINCT t.teacher_id
    FROM student_groups sg
    JOIN teachers t ON sg.teacher_id = t.teacher_id
    WHERE sg.group_name = ?
    """, (group_name,))
    teachers = cursor.fetchall()
    
    for (teacher_id,) in teachers:
        if not is_student_linked(student_id, teacher_id):
            cursor.execute("INSERT OR IGNORE INTO student_teacher (student_id, teacher_id, tokens) VALUES (?, ?, 0)", 
                          (student_id, teacher_id))
        
        cursor.execute("SELECT group_id FROM student_groups WHERE teacher_id = ? AND group_name = ?", 
                      (teacher_id, group_name))
        group = cursor.fetchone()
        if group:
            group_id = group[0]
            cursor.execute("INSERT OR IGNORE INTO group_students (group_id, student_id) VALUES (?, ?)", 
                          (group_id, student_id))
    
    # Начисление зарезервированных жетонов
    cursor.execute("""
    SELECT teacher_id, discipline_id, tokens, notification_message
    FROM reserved_tokens
    WHERE student_id = ?
    """, (student_id,))
    reserved = cursor.fetchall()
    
    for teacher_id, discipline_id, tokens, notification_message in reserved:
        cursor.execute("""
        UPDATE student_teacher 
        SET tokens = tokens + ? 
        WHERE student_id = ? AND teacher_id = ?
        """, (tokens, student_id, teacher_id))
        try:
            await bot.send_message(student_id, notification_message)
            await bot.send_message(teacher_id, f"Студент (ID: {student_id}) зарегистрировался и получил зарезервированные {tokens} жетончиков.")
        except Exception as e:
            logger.error(f"Ошибка отправки уведомления о зарезервированных жетонах: {e}")
    
    cursor.execute("DELETE FROM reserved_tokens WHERE student_id = ?", (student_id,))
    conn.commit()
    
    await state.clear()
    await message.answer("Регистрация завершена. Вы связаны с преподавателями вашей группы.", 
                       reply_markup=ReplyKeyboardRemove())
    await show_student_menu(message)

# Новая функция для ручной проверки и начисления жетонов
@dp.message(F.text == "Проверить сдачу практик")
async def check_practices_completion(message: Message, state: FSMContext):
    if not is_teacher(message.from_user.id):
        await message.answer("Команда доступна только преподавателям.")
        return
    teacher_id = message.from_user.id
    disciplines = get_teacher_disciplines(teacher_id)
    
    if not disciplines:
        await message.answer("У вас нет дисциплин.")
        await state.clear()
        return
    
    keyboard = InlineKeyboardBuilder()
    for discipline in disciplines:
        discipline_id, name, group = discipline
        keyboard.add(InlineKeyboardButton(text=f"{name} ({group})", callback_data=f"check_prac_dis_{discipline_id}"))
    keyboard.adjust(1)
    await message.answer("Выберите дисциплину для проверки сдачи практик:", reply_markup=keyboard.as_markup())

@dp.callback_query(F.data.startswith("check_prac_dis_"))
async def process_check_practices_discipline(callback: types.CallbackQuery, state: FSMContext):
    discipline_id = int(callback.data.split("_")[3])
    teacher_id = callback.from_user.id
    cursor.execute("SELECT name, group_name FROM disciplines WHERE discipline_id = ? AND teacher_id = ?", 
                  (discipline_id, teacher_id))
    discipline = cursor.fetchone()
    
    if not discipline:
        await callback.message.answer("Дисциплина не найдена.")
        await state.clear()
        await callback.answer()
        return
    
    discipline_name, group_name = discipline
    students = get_group_students(group_name, teacher_id)
    
    if not students:
        await callback.message.answer("В этой группе нет студентов.")
        await state.clear()
        await callback.answer()
        return
    
    for student_id, _ in students:
        if check_all_practices_completed(student_id, discipline_id, teacher_id):
            cursor.execute("SELECT tokens FROM reserved_tokens WHERE student_id = ? AND teacher_id = ? AND discipline_id = ?",
                          (student_id, teacher_id, discipline_id))
            if not cursor.fetchone():
                await award_practice_completion(student_id, discipline_id, teacher_id)
    
    await callback.message.answer(f"Проверка сдачи практик по дисциплине '{discipline_name}' завершена. Уведомления и жетончики начислены.")
    await callback.answer()

@dp.message(F.text.regexp(r'^[0-9a-f]{8}$'))
async def process_teacher_token(message: Message, state: FSMContext):
    user_id = message.from_user.id
    token = message.text
    
    cursor.execute("SELECT teacher_id, full_name FROM teachers WHERE token = ?", (token,))
    teacher = cursor.fetchone()
    
    if teacher:
        teacher_id, full_name = teacher
        cursor.execute("INSERT OR REPLACE INTO users (user_id, username, full_name, role) VALUES (?, ?, ?, 'teacher')", 
                      (user_id, message.from_user.username, full_name))
        cursor.execute("UPDATE teachers SET teacher_id = ? WHERE token = ?", (user_id, token))
        cursor.execute("INSERT OR IGNORE INTO admin_teacher (admin_id, teacher_id) VALUES (?, ?)", 
                      (ADMIN_ID, user_id))
        conn.commit()
        await message.answer(f"Добро пожаловать, {full_name}! Вы зарегистрированы как преподаватель.", 
                          reply_markup=ReplyKeyboardRemove())
        await show_teacher_menu(message)
    else:
        await message.answer("Неверный токен. Обратитесь к администратору.")

async def show_admin_menu(message: Message):
    keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="Загрузить список преподавателей")],
            [KeyboardButton(text="Сгенерировать токены преподавателей")],
            [KeyboardButton(text="Просмотреть список преподавателей")],
            [KeyboardButton(text="Просмотреть список студентов")],
            [KeyboardButton(text="Шаблон преподавателей")],
            [KeyboardButton(text="Отправить новость")]
        ],
        resize_keyboard=True
    )
    await message.answer("Меню администратора:", reply_markup=keyboard)

async def show_teacher_menu(message: Message):
    keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="Добавить группу студентов"), KeyboardButton(text="Добавить дисциплину")],
            [KeyboardButton(text="Мои группы"), KeyboardButton(text="Мои дисциплины")],
            [KeyboardButton(text="Создать КТП"), KeyboardButton(text="Просмотреть/удалить КТП")],
            [KeyboardButton(text="Выставить оценки"), KeyboardButton(text="Редактировать оценку")],
            [KeyboardButton(text="Мои студенты"), KeyboardButton(text="Создать ведомость")],
            [KeyboardButton(text="Добавить студента вручную"), KeyboardButton(text="Управление магазином наград")],
            [KeyboardButton(text="Шаблон студентов"), KeyboardButton(text="Настроить жетончики за посещение")],
            [KeyboardButton(text="Проверить сдачу практик")], [KeyboardButton(text="Отправить новость группе")]
        ],
        resize_keyboard=True
    )
    await message.answer("Меню преподавателя:", reply_markup=keyboard)

async def show_student_menu(message: Message):
    keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="Мои оценки"), KeyboardButton(text="Мои преподаватели")],
            [KeyboardButton(text="Экспорт оценок"), KeyboardButton(text="Магазин наград")],
            [KeyboardButton(text="Мои жетончики")], [KeyboardButton(text="Мои практики")]
        ],
        resize_keyboard=True
    )
    await message.answer("Меню студента:", reply_markup=keyboard)

@dp.message(Command("help"))
async def cmd_help(message: Message):
    user_id = message.from_user.id
    user_info = get_user_info(user_id)
    
    if not user_info:
        await message.answer("Доступные команды:\n/start - начать работу\n/help - помощь")
        return
    
    role = user_info[3]
    if role == "admin":
        help_text = """
        Команды администратора:
        /start - открыть меню
        /help - помощь
        /template_teachers - получить шаблон для преподавателей
        /generate_tokens - сгенерировать токены для преподавателей
        /list_teachers - список преподавателей
        /list_students - список студентов
        Загрузить список преподавателей - загрузить Excel файл
        """
    elif role == "teacher":
        help_text = """
        Команды преподавателя:
        /start - открыть меню
        /help - помощь
        /template_students - получить шаблон для студентов
        Добавить группу студентов - загрузить Excel со студентами
        Добавить дисциплину - создать новую дисциплину
        Мои группы - список групп
        Мои дисциплины - список дисциплин
        Создать КТП - создать новый КТП
        Просмотреть/удалить КТП - просмотреть или удалить КТП
        Выставить оценки - начать выставление оценок
        Мои студенты - список студентов
        Создать ведомость - создать Excel ведомость
        Добавить студента вручную - добавить одного студента
        Управление магазином наград - добавить/редактировать/удалить награды
        Настроить жетончики за посещение - установить количество жетончиков за посещение
        """
    else:
        help_text = """
        Команды студента:
        /start - открыть меню
        /help - помощь
        Мои оценки - просмотреть оценки
        Мои преподаватели - список преподавателей
        Экспорт оценок - выгрузить оценки в Excel
        Магазин наград - просмотреть и купить награды
        Мои жетончики - просмотреть баланс жетончиков
        """
    await message.answer(help_text)

@dp.message(F.document, lambda message: is_admin(message.from_user.id))
async def handle_admin_document(message: Message):
    try:
        file_id = message.document.file_id
        file = await bot.get_file(file_id)
        file_path = file.file_path
        await bot.download_file(file_path, "uploaded_teachers.xlsx")
        
        df = pd.read_excel("uploaded_teachers.xlsx")
        if "ФИО преподавателя" not in df.columns:
            await message.answer("Ожидается колонка 'ФИО преподавателя'.")
            return
        
        added = 0
        for _, row in df.iterrows():
            full_name = str(row["ФИО преподавателя"]).strip()
            if not full_name or pd.isna(full_name):
                continue
            cursor.execute("SELECT 1 FROM teachers WHERE full_name = ?", (full_name,))
            if not cursor.fetchone():
                token = generate_token()
                cursor.execute("INSERT INTO teachers (full_name, token, tokens_per_attendance) VALUES (?, ?, 1)", (full_name, token))
                teacher_id = cursor.lastrowid
                cursor.execute("INSERT OR IGNORE INTO admin_teacher (admin_id, teacher_id) VALUES (?, ?)", 
                             (message.from_user.id, teacher_id))
                added += 1
        
        conn.commit()
        await message.answer(f"Добавлено {added} преподавателей.")
        os.remove("uploaded_teachers.xlsx")
    except Exception as e:
        logger.error(f"Error processing admin document: {e}")
        await message.answer("Ошибка при обработке файла.")

@dp.message(F.text == "Добавить группу студентов")
async def add_student_group_start(message: Message, state: FSMContext):
    if not is_teacher(message.from_user.id):
        await message.answer("Команда доступна только преподавателям.")
        return
    await state.set_state(TeacherForm.waiting_for_student_list)
    await message.answer("Загрузите Excel файл со списком студентов (колонки: ФИО студента, Группа).")

@dp.message(TeacherForm.waiting_for_student_list, F.document)
async def handle_teacher_document(message: Message, state: FSMContext):
    try:
        file_id = message.document.file_id
        file = await bot.get_file(file_id)
        file_path = file.file_path
        await bot.download_file(file_path, "uploaded_students.xlsx")
        
        df = pd.read_excel("uploaded_students.xlsx")
        if "ФИО студента" not in df.columns or "Группа" not in df.columns:
            await message.answer("Ожидаются колонки 'ФИО студента' и 'Группа'.")
            os.remove("uploaded_students.xlsx")
            return
        
        await state.update_data(student_df=df.to_dict('records'))
        await state.set_state(TeacherForm.waiting_for_group_name)
        await message.answer("Введите название группы для этих студентов:")
        os.remove("uploaded_students.xlsx")
    except Exception as e:
        logger.error(f"Error processing teacher document: {e}")
        await message.answer("Ошибка при обработке файла.")
        if os.path.exists("uploaded_students.xlsx"):
            os.remove("uploaded_students.xlsx")

@dp.message(TeacherForm.waiting_for_group_name)
async def process_group_name(message: Message, state: FSMContext):
    teacher_id = message.from_user.id
    group_name = message.text.strip()
    data = await state.get_data()
    students = data.get("student_df", [])
    
    cursor.execute("INSERT OR IGNORE INTO student_groups (teacher_id, group_name) VALUES (?, ?)", (teacher_id, group_name))
    cursor.execute("SELECT group_id FROM student_groups WHERE teacher_id = ? AND group_name = ?", (teacher_id, group_name))
    group_id = cursor.fetchone()[0]
    
    added = 0
    for student in students:
        full_name = str(student.get("ФИО студента")).strip()
        student_group = str(student.get("Группа", group_name)).strip()
        if not full_name or pd.isna(full_name):
            continue
        
        cursor.execute("SELECT student_id FROM students WHERE full_name = ? AND group_name = ?", (full_name, student_group))
        existing_student = cursor.fetchone()
        
        if existing_student:
            student_id = existing_student[0]
        else:
            cursor.execute("INSERT INTO students (full_name, group_name) VALUES (?, ?)", 
                          (full_name, student_group))
            student_id = cursor.lastrowid
            cursor.execute("INSERT OR IGNORE INTO users (user_id, full_name, role, group_name) VALUES (?, ?, 'student', ?)", 
                          (student_id, full_name, student_group))
            cursor.execute("INSERT OR IGNORE INTO admin_student (admin_id, student_id) VALUES (?, ?)", 
                          (ADMIN_ID, student_id))
            added += 1
        
        cursor.execute("INSERT OR IGNORE INTO group_students (group_id, student_id) VALUES (?, ?)", (group_id, student_id))
        if not is_student_linked(student_id, teacher_id):
            cursor.execute("INSERT OR IGNORE INTO student_teacher (student_id, teacher_id, tokens) VALUES (?, ?, 0)", 
                          (student_id, teacher_id))
    
    conn.commit()
    await state.clear()
    await message.answer(f"Добавлено {added} новых студентов в группу '{group_name}'.")

@dp.message(F.text == "Добавить студента вручную")
async def add_student_manually_start(message: Message, state: FSMContext):
    if not is_teacher(message.from_user.id):
        await message.answer("Команда доступна только преподавателям.")
        return
    await state.set_state(TeacherForm.waiting_for_student_name)
    await message.answer("Введите ФИО студента:")

@dp.message(TeacherForm.waiting_for_student_name)
async def process_student_name_manual(message: Message, state: FSMContext):
    await state.update_data(student_name=message.text.strip())
    teacher_id = message.from_user.id
    groups = get_teacher_groups(teacher_id)
    
    if not groups:
        await message.answer("У вас нет групп. Сначала добавьте группу.")
        await state.clear()
        return
    
    keyboard = InlineKeyboardBuilder()
    for group in groups:
        keyboard.add(InlineKeyboardButton(text=group, callback_data=f"manual_group_{group}"))
    keyboard.adjust(1)
    await state.set_state(TeacherForm.waiting_for_student_group)
    await message.answer("Выберите группу для студента:", reply_markup=keyboard.as_markup())

@dp.callback_query(F.data.startswith("manual_group_"))
async def process_student_group_manual(callback: types.CallbackQuery, state: FSMContext):
    teacher_id = callback.from_user.id
    group_name = callback.data.split("_")[2]
    data = await state.get_data()
    full_name = data.get("student_name")
    
    cursor.execute("SELECT student_id FROM students WHERE full_name = ? AND group_name = ?", (full_name, group_name))
    existing_student = cursor.fetchone()
    
    if existing_student:
        student_id = existing_student[0]
    else:
        cursor.execute("INSERT INTO students (full_name, group_name) VALUES (?, ?)", (full_name, group_name))
        student_id = cursor.lastrowid
        cursor.execute("INSERT OR IGNORE INTO users (user_id, full_name, role, group_name) VALUES (?, ?, 'student', ?)", 
                      (student_id, full_name, group_name))
        cursor.execute("INSERT OR IGNORE INTO admin_student (admin_id, student_id) VALUES (?, ?)", 
                      (ADMIN_ID, student_id))
    
    cursor.execute("SELECT group_id FROM student_groups WHERE teacher_id = ? AND group_name = ?", 
                  (teacher_id, group_name))
    group = cursor.fetchone()
    
    if group:
        group_id = group[0]
        cursor.execute("INSERT OR IGNORE INTO group_students (group_id, student_id) VALUES (?, ?)", (group_id, student_id))
    
    if not is_student_linked(student_id, teacher_id):
        cursor.execute("INSERT OR IGNORE INTO student_teacher (student_id, teacher_id, tokens) VALUES (?, ?, 0)", 
                      (student_id, teacher_id))
    
    conn.commit()
    await state.clear()
    await callback.message.answer(f"Студент {full_name} добавлен в группу {group_name}.")
    await callback.answer()

@dp.message(F.text == "Добавить дисциплину")
async def add_discipline_start(message: Message, state: FSMContext):
    if not is_teacher(message.from_user.id):
        await message.answer("Команда доступна только преподавателям.")
        return
    await state.set_state(TeacherForm.waiting_for_discipline_name)
    await message.answer("Введите название дисциплины:")

@dp.message(TeacherForm.waiting_for_discipline_name)
async def process_discipline_name(message: Message, state: FSMContext):
    teacher_id = message.from_user.id
    groups = get_teacher_groups(teacher_id)
    
    if not groups:
        await message.answer("У вас нет групп. Сначала добавьте группу.")
        await state.clear()
        return
    
    await state.update_data(discipline_name=message.text.strip())
    await state.set_state(TeacherForm.waiting_for_required_practices)
    await message.answer("Введите количество необходимых практик для дисциплины (целое число ≥ 0):")

@dp.message(TeacherForm.waiting_for_required_practices)
async def process_required_practices(message: Message, state: FSMContext):
    try:
        required_practices = int(message.text)
        if required_practices < 0:
            await message.answer("Количество практик не может быть отрицательным. Попробуйте снова.")
            return
    except ValueError:
        await message.answer("Введите целое число. Попробуйте снова.")
        return
    
    await state.update_data(required_practices=required_practices)
    teacher_id = message.from_user.id
    groups = get_teacher_groups(teacher_id)
    
    keyboard = InlineKeyboardBuilder()
    for group in groups:
        keyboard.add(InlineKeyboardButton(text=group, callback_data=f"dis_group_{group}"))
    keyboard.adjust(1)
    await message.answer("Выберите группу для дисциплины:", reply_markup=keyboard.as_markup())

@dp.callback_query(F.data.startswith("dis_group_"))
async def process_discipline_group(callback: types.CallbackQuery, state: FSMContext):
    group_name = callback.data.split("_")[2]
    teacher_id = callback.from_user.id
    data = await state.get_data()
    
    discipline_name = data.get("discipline_name")
    required_practices = data.get("required_practices")
    
    cursor.execute("""
    INSERT INTO disciplines (teacher_id, name, group_name, required_practices)
    VALUES (?, ?, ?, ?)
    """, (teacher_id, discipline_name, group_name, required_practices))
    conn.commit()
    
    await state.clear()
    await callback.message.answer(f"Дисциплина '{discipline_name}' с {required_practices} практиками добавлена для группы '{group_name}'.")
    await callback.answer()

@dp.message(F.text == "Создать КТП")
async def create_ktp_start(message: Message, state: FSMContext):
    if not is_teacher(message.from_user.id):
        await message.answer("Команда доступна только преподавателями.")
        return
    teacher_id = message.from_user.id
    disciplines = get_teacher_disciplines(teacher_id)
    
    if not disciplines:
        await message.answer("У вас нет дисциплин.")
        await state.clear()
        return
    
    keyboard = InlineKeyboardBuilder()
    for discipline in disciplines:
        discipline_id, name, group = discipline
        keyboard.add(InlineKeyboardButton(text=name, callback_data=f"ktp_dis_{discipline_id}"))
    keyboard.adjust(1)
    await state.set_state(TeacherForm.waiting_for_ktp_discipline)
    await message.answer("Выберите дисциплину для КТП:", reply_markup=keyboard.as_markup())

@dp.callback_query(F.data.startswith("ktp_dis_"))
async def process_ktp_discipline(callback: types.CallbackQuery, state: FSMContext):
    discipline_id = int(callback.data.split("_")[2])
    teacher_id = callback.from_user.id
    cursor.execute("SELECT name FROM disciplines WHERE discipline_id = ? AND teacher_id = ?", 
                  (discipline_id, teacher_id))
    discipline = cursor.fetchone()
    
    if not discipline:
        await callback.message.answer("Дисциплина не найдена.")
        await state.clear()
        await callback.answer()
        return
    
    await state.update_data(discipline_id=discipline_id, discipline_name=discipline[0])
    keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="Лекция"), KeyboardButton(text="Практика")]
        ],
        resize_keyboard=True
    )
    await state.set_state(TeacherForm.waiting_for_ktp_type)
    await callback.message.answer("Выберите тип КТП:", reply_markup=keyboard)
    await callback.answer()

@dp.message(TeacherForm.waiting_for_ktp_type, F.text.in_(["Лекция", "Практика"]))
async def process_ktp_type(message: Message, state: FSMContext):
    ktp_type = "lecture" if message.text == "Лекция" else "practice"
    await state.update_data(ktp_type=ktp_type)
    
    if ktp_type == "lecture":
        await state.set_state(TeacherForm.waiting_for_lecture_topic)
        await message.answer("Введите тему лекции:", reply_markup=ReplyKeyboardRemove())
    else:
        keyboard = ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text=str(i)) for i in range(1, 6)],
                      [KeyboardButton(text=str(i)) for i in range(6, 11)],
                      [KeyboardButton(text=str(i)) for i in range(11, 16)],
                      [KeyboardButton(text=str(i)) for i in range(16, 21)]],
            resize_keyboard=True
        )
        await state.set_state(TeacherForm.waiting_for_practice_number)
        await message.answer("Выберите номер практики (1-20):", reply_markup=keyboard)

@dp.message(TeacherForm.waiting_for_lecture_topic)
async def process_lecture_topic(message: Message, state: FSMContext):
    description = message.text.strip()
    await state.update_data(description=description)
    await state.set_state(TeacherForm.waiting_for_lecture_homework)
    await message.answer("Введите домашнее задание для лекции (или 'нет', если отсутствует):", reply_markup=ReplyKeyboardRemove())

@dp.message(TeacherForm.waiting_for_lecture_homework)
async def process_lecture_homework(message: Message, state: FSMContext):
    homework = message.text.strip() if message.text.strip().lower() != 'нет' else None
    teacher_id = message.from_user.id
    data = await state.get_data()
    discipline_id = data.get("discipline_id")
    ktp_type = data.get("ktp_type")
    description = data.get("description")
    
    # Находим все группы, связанные с этой дисциплиной
    cursor.execute("SELECT group_name FROM disciplines WHERE discipline_id = ?", (discipline_id,))
    groups = [row[0] for row in cursor.fetchall()]
    
    for group_name in groups:
        cursor.execute("""
        INSERT INTO ktp (teacher_id, discipline_id, group_name, type, description, practice_number, homework)
        VALUES (?, ?, ?, ?, ?, NULL, ?)
        """, (teacher_id, discipline_id, group_name, ktp_type, description, homework))
    
    conn.commit()
    
    await state.clear()
    await message.answer(f"КТП для лекции создан.", reply_markup=ReplyKeyboardRemove())
    await show_teacher_menu(message)

@dp.message(TeacherForm.waiting_for_practice_number)
async def process_practice_number(message: Message, state: FSMContext):
    try:
        practice_number = int(message.text)
        if not 1 <= practice_number <= 20:
            await message.answer("Номер практики должен быть от 1 до 20. Попробуйте снова.")
            return
    except ValueError:
        await message.answer("Введите число от 1 до 20. Попробуйте снова.")
        return
    
    await state.update_data(practice_number=practice_number)
    await state.set_state(TeacherForm.waiting_for_practice_topic)
    await message.answer("Введите тему практики:", reply_markup=ReplyKeyboardRemove())

@dp.message(TeacherForm.waiting_for_practice_topic)
async def process_practice_topic(message: Message, state: FSMContext):
    description = message.text.strip()
    await state.update_data(description=description)
    await state.set_state(TeacherForm.waiting_for_practice_homework)
    await message.answer("Введите домашнее задание для практики (или 'нет', если отсутствует):", reply_markup=ReplyKeyboardRemove())

@dp.message(TeacherForm.waiting_for_practice_homework)
async def process_practice_homework(message: Message, state: FSMContext):
    homework = message.text.strip() if message.text.strip().lower() != 'нет' else None
    teacher_id = message.from_user.id
    data = await state.get_data()
    discipline_id = data.get("discipline_id")
    ktp_type = data.get("ktp_type")
    practice_number = data.get("practice_number")
    description = data.get("description")
    
    # Находим все группы, связанные с этой дисциплиной
    cursor.execute("SELECT group_name FROM disciplines WHERE discipline_id = ?", (discipline_id,))
    groups = [row[0] for row in cursor.fetchall()]
    
    for group_name in groups:
        cursor.execute("""
        INSERT INTO ktp (teacher_id, discipline_id, group_name, type, description, practice_number, homework)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (teacher_id, discipline_id, group_name, ktp_type, description, practice_number, homework))
    
    conn.commit()
    
    await state.clear()
    await message.answer(f"КТП для практики #{practice_number} создан.", reply_markup=ReplyKeyboardRemove())
    await show_teacher_menu(message)

@dp.message(F.text == "Просмотреть/удалить КТП")
async def view_delete_ktp_start(message: Message, state: FSMContext):
    if not is_teacher(message.from_user.id):
        await message.answer("Команда доступна только преподавателями.")
        return
    teacher_id = message.from_user.id
    disciplines = get_teacher_disciplines(teacher_id)
    
    if not disciplines:
        await message.answer("У вас нет дисциплин.")
        return
    
    keyboard = InlineKeyboardBuilder()
    for discipline in disciplines:
        discipline_id, name, group = discipline
        keyboard.add(InlineKeyboardButton(text=name, callback_data=f"view_ktp_dis_{discipline_id}"))
    keyboard.adjust(1)
    await state.set_state(TeacherForm.waiting_for_view_ktp_discipline)
    await message.answer("Выберите дисциплину для просмотра КТП:", reply_markup=keyboard.as_markup())

@dp.callback_query(F.data.startswith("view_ktp_dis_"))
async def view_ktp_by_discipline(callback: types.CallbackQuery, state: FSMContext):
    discipline_id = int(callback.data.split("_")[3])
    teacher_id = callback.from_user.id
    cursor.execute("SELECT name FROM disciplines WHERE discipline_id = ? AND teacher_id = ?", 
                  (discipline_id, teacher_id))
    discipline = cursor.fetchone()
    
    if not discipline:
        await callback.message.answer("Дисциплина не найдена.")
        await callback.answer()
        return
    
    ktps = get_ktp_by_discipline_and_type(teacher_id, discipline_id, 'lecture') + get_ktp_by_discipline_and_type(teacher_id, discipline_id, 'practice')
    
    if not ktps:
        await callback.message.answer(f"Нет КТП для дисциплины '{discipline[0]}'.")
        await callback.answer()
        return
    
    response = f"КТП для дисциплины '{discipline[0]}':\n"
    keyboard = InlineKeyboardBuilder()
    for ktp_id, group_name, ktp_type, description, practice_number in ktps:
        ktp_type_str = "Лекция" if ktp_type == "lecture" else f"Практика #{practice_number}"
        response += f"- {group_name}, {ktp_type_str}: {description}\n"
        keyboard.add(InlineKeyboardButton(
            text=f"Удалить: {ktp_type_str} ({group_name})",
            callback_data=f"delete_ktp_{ktp_id}"
        ))
    keyboard.adjust(1)
    
    await callback.message.answer(response, reply_markup=keyboard.as_markup())
    await callback.answer()

@dp.callback_query(F.data.startswith("delete_ktp_"))
async def delete_ktp(callback: types.CallbackQuery):
    ktp_id = int(callback.data.split("_")[2])
    cursor.execute("SELECT discipline_id, group_name, type, description, practice_number FROM ktp WHERE ktp_id = ?", (ktp_id,))
    ktp = cursor.fetchone()
    
    if not ktp:
        await callback.message.answer("КТП не найден.")
        await callback.answer()
        return
    
    discipline_id, group_name, ktp_type, description, practice_number = ktp
    cursor.execute("SELECT name FROM disciplines WHERE discipline_id = ?", (discipline_id,))
    discipline_name = cursor.fetchone()[0]
    
    cursor.execute("DELETE FROM grades WHERE ktp_id = ?", (ktp_id,))
    cursor.execute("DELETE FROM ktp WHERE ktp_id = ?", (ktp_id,))
    conn.commit()
    
    ktp_type_str = "Лекция" if ktp_type == "lecture" else f"Практика #{practice_number}"
    await callback.message.answer(f"КТП для '{discipline_name}' ({ktp_type_str}, {group_name}) удален.")
    await callback.answer()

@dp.message(F.text == "Настроить жетончики за посещение")
async def set_tokens_per_attendance_start(message: Message, state: FSMContext):
    if not is_teacher(message.from_user.id):
        await message.answer("Команда доступна только преподавателями.")
        return
    teacher_id = message.from_user.id
    cursor.execute("SELECT tokens_per_attendance FROM teachers WHERE teacher_id = ?", (teacher_id,))
    current_tokens = cursor.fetchone()[0]
    
    await state.set_state(TeacherForm.waiting_for_tokens_per_attendance)
    keyboard = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text=str(current_tokens))]],
        resize_keyboard=True
    )
    await message.answer(f"Текущее количество жетончиков за посещение: {current_tokens}\nВведите новое количество:", reply_markup=keyboard)

@dp.message(TeacherForm.waiting_for_tokens_per_attendance)
async def process_tokens_per_attendance(message: Message, state: FSMContext):
    try:
        tokens = int(message.text)
        if tokens < 0:
            await message.answer("Количество жетончиков не может быть отрицательным. Попробуйте снова.")
            return
    except ValueError:
        await message.answer("Введите целое число. Попробуйте снова.")
        return
    
    teacher_id = message.from_user.id
    cursor.execute("UPDATE teachers SET tokens_per_attendance = ? WHERE teacher_id = ?", (tokens, teacher_id))
    conn.commit()
    
    await state.clear()
    await message.answer(f"Количество жетончиков за посещение установлено: {tokens}.", reply_markup=ReplyKeyboardRemove())
    await show_teacher_menu(message)

@dp.message(F.text == "Управление магазином наград")
async def manage_rewards_start(message: Message, state: FSMContext):
    if not is_teacher(message.from_user.id):
        await message.answer("Команда доступна только преподавателями.")
        return
    teacher_id = message.from_user.id
    disciplines = get_teacher_disciplines(teacher_id)
    
    if not disciplines:
        await message.answer("У вас нет дисциплин. Сначала добавьте дисциплину.")
        return
    
    keyboard = InlineKeyboardBuilder()
    keyboard.add(InlineKeyboardButton(text="Добавить награду", callback_data="add_reward"))
    keyboard.add(InlineKeyboardButton(text="Просмотреть награды", callback_data="view_rewards"))
    keyboard.adjust(1)
    await message.answer("Выберите действие:", reply_markup=keyboard.as_markup())

@dp.callback_query(F.data == "add_reward")
async def add_reward_start(callback: types.CallbackQuery, state: FSMContext):
    teacher_id = callback.from_user.id
    disciplines = get_teacher_disciplines(teacher_id)
    
    if not disciplines:
        await callback.message.answer("У вас нет дисциплин.")
        await callback.answer()
        return
    
    keyboard = InlineKeyboardBuilder()
    for discipline in disciplines:
        discipline_id, name, group = discipline
        keyboard.add(InlineKeyboardButton(text=f"{name} ({group})", callback_data=f"reward_dis_{discipline_id}"))
    keyboard.adjust(1)
    await state.set_state(TeacherForm.waiting_for_reward_name)
    await callback.message.answer("Выберите дисциплину для награды:", reply_markup=keyboard.as_markup())
    await callback.answer()

@dp.callback_query(F.data.startswith("reward_dis_"))
async def select_discipline_for_reward(callback: types.CallbackQuery, state: FSMContext):
    discipline_id = int(callback.data.split("_")[2])
    cursor.execute("SELECT name, group_name FROM disciplines WHERE discipline_id = ?", (discipline_id,))
    discipline = cursor.fetchone()
    
    if not discipline:
        await callback.message.answer("Дисциплина не найдена.")
        await callback.answer()
        return
    
    await state.update_data(discipline_id=discipline_id, discipline_name=discipline[0])
    await callback.message.answer("Введите название награды:")
    await callback.answer()

@dp.message(TeacherForm.waiting_for_reward_name)
async def process_reward_name(message: Message, state: FSMContext):
    await state.update_data(reward_name=message.text.strip())
    await state.set_state(TeacherForm.waiting_for_reward_description)
    await message.answer("Введите описание награды:")

@dp.message(TeacherForm.waiting_for_reward_description)
async def process_reward_description(message: Message, state: FSMContext):
    await state.update_data(reward_description=message.text.strip())
    await state.set_state(TeacherForm.waiting_for_reward_price)
    await message.answer("Введите цену награды в жетончиках (целое число):")

@dp.message(TeacherForm.waiting_for_reward_price)
async def process_reward_price(message: Message, state: FSMContext):
    try:
        price = int(message.text)
        if price <= 0:
            await message.answer("Цена должна быть положительным числом. Попробуйте снова.")
            return
    except ValueError:
        await message.answer("Введите целое число. Попробуйте снова.")
        return
    
    teacher_id = message.from_user.id
    data = await state.get_data()
    discipline_id = data.get("discipline_id")
    reward_name = data.get("reward_name")
    reward_description = data.get("reward_description")
    
    cursor.execute("""
    INSERT INTO rewards (teacher_id, discipline_id, name, description, price, is_enabled)
    VALUES (?, ?, ?, ?, ?, 1)
    """, (teacher_id, discipline_id, reward_name, reward_description, price))
    conn.commit()
    
    await state.clear()
    await message.answer(f"Награда '{reward_name}' добавлена для дисциплины '{data.get('discipline_name')}'.")
    await show_teacher_menu(message)

@dp.callback_query(F.data == "view_rewards")
async def view_teacher_rewards(callback: types.CallbackQuery):
    teacher_id = callback.from_user.id
    rewards = get_teacher_rewards(teacher_id)
    
    if not rewards:
        await callback.message.answer("У вас нет наград.")
        await callback.answer()
        return
    
    response = "Ваши награды:\n"
    keyboard = InlineKeyboardBuilder()
    for reward_id, name, description, price, is_enabled, discipline_name in rewards:
        status = "Включена" if is_enabled else "Отключена"
        response += f"- {name} ({discipline_name}): {description}, {price} токенов ({status})\n"
        keyboard.add(InlineKeyboardButton(text=f"Редактировать цену: {name}", callback_data=f"edit_reward_price_{reward_id}"))
        keyboard.add(InlineKeyboardButton(text=f"{'Отключить' if is_enabled else 'Включить'}: {name}", 
                                       callback_data=f"toggle_reward_{reward_id}"))
        keyboard.add(InlineKeyboardButton(text=f"Удалить: {name}", callback_data=f"delete_reward_{reward_id}"))
    keyboard.adjust(1)
    await callback.message.answer(response, reply_markup=keyboard.as_markup())
    await callback.answer()

@dp.callback_query(F.data.startswith("edit_reward_price_"))
async def edit_reward_price_start(callback: types.CallbackQuery, state: FSMContext):
    reward_id = int(callback.data.split("_")[3])
    cursor.execute("SELECT name, price, discipline_id FROM rewards WHERE reward_id = ?", (reward_id,))
    reward = cursor.fetchone()
    
    if not reward:
        await callback.message.answer("Награда не найдена.")
        await callback.answer()
        return
    
    await state.set_state(TeacherForm.waiting_for_edit_reward_price)
    await state.update_data(reward_id=reward_id, reward_name=reward[0])
    await callback.message.answer(f"Введите новую цену для награды '{reward[0]}' (текущая: {reward[1]} токенов):",
                               reply_markup=ReplyKeyboardMarkup(
                                   keyboard=[[KeyboardButton(text=str(reward[1]))]], 
                                   resize_keyboard=True))
    await callback.answer()

@dp.message(TeacherForm.waiting_for_edit_reward_price)
async def process_edit_reward_price(message: Message, state: FSMContext):
    try:
        price = int(message.text)
        if price <= 0:
            await message.answer("Цена должна быть положительным числом. Попробуйте снова.")
            return
    except ValueError:
        await message.answer("Введите целое число. Попробуйте снова.")
        return
    
    data = await state.get_data()
    reward_id = data.get("reward_id")
    reward_name = data.get("reward_name")
    
    cursor.execute("UPDATE rewards SET price = ? WHERE reward_id = ?", (price, reward_id))
    conn.commit()
    
    await state.clear()
    await message.answer(f"Цена награды '{reward_name}' обновлена до {price} токенов.",
                       reply_markup=ReplyKeyboardRemove())
    await show_teacher_menu(message)

@dp.callback_query(F.data.startswith("toggle_reward_"))
async def toggle_reward_status(callback: types.CallbackQuery):
    reward_id = int(callback.data.split("_")[2])
    cursor.execute("SELECT name, is_enabled FROM rewards WHERE reward_id = ?", (reward_id,))
    reward = cursor.fetchone()
    
    if not reward:
        await callback.message.answer("Награда не найдена.")
        await callback.answer()
        return
    
    new_status = 0 if reward[1] else 1
    cursor.execute("UPDATE rewards SET is_enabled = ? WHERE reward_id = ?", (new_status, reward_id))
    conn.commit()
    
    status_text = "включена" if new_status else "отключена"
    await callback.message.answer(f"Награда '{reward[0]}' {status_text}.")
    await callback.answer()

@dp.callback_query(F.data.startswith("delete_reward_"))
async def delete_reward(callback: types.CallbackQuery):
    reward_id = int(callback.data.split("_")[2])
    cursor.execute("SELECT name FROM rewards WHERE reward_id = ?", (reward_id,))
    reward = cursor.fetchone()
    
    if not reward:
        await callback.message.answer("Награда не найдена.")
        await callback.answer()
        return
    
    cursor.execute("DELETE FROM purchased_rewards WHERE reward_id = ?", (reward_id,))
    cursor.execute("DELETE FROM rewards WHERE reward_id = ?", (reward_id,))
    conn.commit()
    
    await callback.message.answer(f"Награда '{reward[0]}' удалена.")
    await callback.answer()

@dp.message(F.text == "Мои дисциплины")
async def show_teacher_disciplines(message: Message):
    teacher_id = message.from_user.id
    disciplines = get_teacher_disciplines(teacher_id)
    
    if not disciplines:
        await message.answer("У вас нет дисциплин.")
        return
    
    response = "Ваши дисциплины:\n"
    for discipline in disciplines:
        discipline_id, name, group = discipline
        response += f"- {name} (группа: {group})\n"
    
    keyboard = InlineKeyboardBuilder()
    for discipline in disciplines:
        discipline_id, name, group = discipline
        keyboard.add(InlineKeyboardButton(text=f"Удалить: {name} ({group})", 
                                       callback_data=f"del_dis_{discipline_id}"))
    keyboard.adjust(1)
    await message.answer(response, reply_markup=keyboard.as_markup())

@dp.callback_query(F.data.startswith("del_dis_"))
async def delete_discipline(callback: types.CallbackQuery):
    discipline_id = int(callback.data.split("_")[2])
    cursor.execute("SELECT name, group_name FROM disciplines WHERE discipline_id = ?", (discipline_id,))
    discipline = cursor.fetchone()
    
    if discipline:
        cursor.execute("DELETE FROM grades WHERE discipline_id = ?", (discipline_id,))
        cursor.execute("DELETE FROM rewards WHERE discipline_id = ?", (discipline_id,))
        cursor.execute("DELETE FROM ktp WHERE discipline_id = ?", (discipline_id,))
        cursor.execute("DELETE FROM disciplines WHERE discipline_id = ?", (discipline_id,))
        conn.commit()
        await callback.message.answer(f"Дисциплина '{discipline[0]}' ({discipline[1]}) удалена.")
    else:
        await callback.message.answer("Дисциплина не найдена.")
    
    await callback.answer()

@dp.message(F.text == "Мои группы")
async def show_teacher_groups(message: Message):
    teacher_id = message.from_user.id
    groups = get_teacher_groups(teacher_id)
    
    if not groups:
        await message.answer("У вас нет групп.")
        return
    
    response = "Ваши группы:\n"
    for group in groups:
        response += f"- {group}\n"
    
    keyboard = InlineKeyboardBuilder()
    for group in groups:
        keyboard.add(InlineKeyboardButton(text=f"Удалить: {group}", callback_data=f"del_group_{group}"))
    keyboard.adjust(1)
    await message.answer(response, reply_markup=keyboard.as_markup())

@dp.callback_query(F.data.startswith("del_group_"))
async def delete_group(callback: types.CallbackQuery):
    group_name = callback.data.split("_")[2]
    teacher_id = callback.from_user.id
    
    cursor.execute("SELECT group_id FROM student_groups WHERE teacher_id = ? AND group_name = ?", 
                  (teacher_id, group_name))
    group = cursor.fetchone()
    
    if group:
        group_id = group[0]
        cursor.execute("DELETE FROM group_students WHERE group_id = ?", (group_id,))
        cursor.execute("DELETE FROM student_groups WHERE group_id = ?", (group_id,))
        cursor.execute("DELETE FROM disciplines WHERE teacher_id = ? AND group_name = ?", 
                      (teacher_id, group_name))
        cursor.execute("DELETE FROM ktp WHERE teacher_id = ? AND group_name = ?", 
                      (teacher_id, group_name))
        cursor.execute("DELETE FROM rewards WHERE teacher_id = ? AND discipline_id IN (SELECT discipline_id FROM disciplines WHERE group_name = ?)", 
                      (teacher_id, group_name))
        conn.commit()
        await callback.message.answer(f"Группа '{group_name}' удалена.")
    else:
        await callback.message.answer("Группа не найдена.")
    
    await callback.answer()

@dp.message(F.text == "Выставить оценки")
async def start_grading(message: Message, state: FSMContext):
    if not is_teacher(message.from_user.id):
        await message.answer("Команда доступна только преподавателями.")
        return
    teacher_id = message.from_user.id
    disciplines = get_teacher_disciplines(teacher_id)
    
    if not disciplines:
        await message.answer("У вас нет дисциплин.")
        await state.clear()
        return
    
    keyboard = InlineKeyboardBuilder()
    for discipline in disciplines:
        discipline_id, name, group = discipline
        keyboard.add(InlineKeyboardButton(text=f"{name} ({group})", callback_data=f"grade_dis_{discipline_id}"))
    keyboard.adjust(1)
    await state.set_state(GradeForm.selecting_discipline)
    await message.answer("Выберите дисциплину для выставления оценок:", reply_markup=keyboard.as_markup())

@dp.callback_query(F.data.startswith("grade_dis_"))
async def select_discipline_for_grading(callback: types.CallbackQuery, state: FSMContext):
    discipline_id = int(callback.data.split("_")[2])
    teacher_id = callback.from_user.id
    cursor.execute("SELECT name, group_name FROM disciplines WHERE discipline_id = ? AND teacher_id = ?", 
                  (discipline_id, teacher_id))
    discipline = cursor.fetchone()
    
    if not discipline:
        await callback.message.answer("Дисциплина не найдена.")
        await state.clear()
        await callback.answer()
        return
    
    await state.update_data(discipline_id=discipline_id, discipline_name=discipline[0], group_name=discipline[1])
    keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="Лекция"), KeyboardButton(text="Практика")]
        ],
        resize_keyboard=True
    )
    await state.set_state(GradeForm.selecting_ktp_type)
    await callback.message.answer("Выберите тип КТП:", reply_markup=keyboard)
    await callback.answer()

# Обновление функции выбора КТП для оценок
@dp.message(GradeForm.selecting_ktp_type, F.text.in_(["Лекция", "Практика"]))
async def process_ktp_type_for_grading(message: Message, state: FSMContext):
    ktp_type = "lecture" if message.text == "Лекция" else "practice"
    teacher_id = message.from_user.id
    data = await state.get_data()
    discipline_id = data.get("discipline_id")
    discipline_name = data.get("discipline_name")
    group_name = data.get("group_name")
    
    ktps = get_ktp_by_discipline_and_type(teacher_id, discipline_id, ktp_type)
    
    if not ktps:
        await message.answer(f"Нет КТП типа '{message.text}' для дисциплины '{discipline_name}'.")
        await state.clear()
        await show_teacher_menu(message)
        return
    
    if ktp_type == "practice":
        cursor.execute("SELECT required_practices FROM disciplines WHERE discipline_id = ?", (discipline_id,))
        required_practices = cursor.fetchone()[0]
        response = f"Необходимое количество практик для дисциплины '{discipline_name}': {required_practices}\n"
        for ktp_id, group_name, ktp_type, description, practice_number in ktps:
            cursor.execute("""
            SELECT COUNT(DISTINCT g.student_id)
            FROM grades g
            JOIN ktp k ON g.ktp_id = k.ktp_id
            WHERE k.practice_number = ? AND g.discipline_id = ? AND g.grade >= 1 AND g.grade != -1
            """, (practice_number, discipline_id))
            completed_count = cursor.fetchone()[0]
            response += f"Практика #{practice_number} ({description}): Сдано студентами: {completed_count}\n"
        await message.answer(response)
    
    await state.update_data(ktp_type=ktp_type)
    keyboard = InlineKeyboardBuilder()
    for ktp_id, group_name, ktp_type, description, practice_number in ktps:
        ktp_type_str = "Лекция" if ktp_type == "lecture" else f"Практика #{practice_number}"
        keyboard.add(InlineKeyboardButton(
            text=f"{description} ({group_name})",
            callback_data=f"grade_ktp_{ktp_id}_{discipline_id}"
        ))
    keyboard.adjust(1)
    await state.set_state(GradeForm.selecting_ktp)
    await message.answer(f"Выберите КТП для дисциплины '{discipline_name}':", reply_markup=keyboard.as_markup())

@dp.callback_query(F.data.startswith("grade_ktp_"))
async def select_ktp_for_grading(callback: types.CallbackQuery, state: FSMContext):
    ktp_id = int(callback.data.split("_")[2])
    discipline_id = int(callback.data.split("_")[3])
    teacher_id = callback.from_user.id
    
    cursor.execute("SELECT name, group_name FROM disciplines WHERE discipline_id = ?", (discipline_id,))
    discipline = cursor.fetchone()
    if not discipline:
        await callback.message.answer("Дисциплина не найдена.")
        await state.clear()
        await callback.answer()
        return
    
    discipline_name, group_name = discipline
    students = get_group_students(group_name, teacher_id)
    
    if not students:
        await callback.message.answer("В этой группе нет студентов.")
        await state.clear()
        await callback.answer()
        return
    
    await state.update_data(
        ktp_id=ktp_id,
        discipline_id=discipline_id,
        discipline_name=discipline_name,
        group_name=group_name,
        students=students
    )
    
    today = datetime.now().strftime("%d-%m-%Y")
    keyboard = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text=today)]],
        resize_keyboard=True
    )
    await state.set_state(GradeForm.selecting_date)
    await callback.message.answer(f"Введите дату для выставления оценок (дд-мм-гггг, по умолчанию {today}):", reply_markup=keyboard)
    await callback.answer()

@dp.message(GradeForm.selecting_date)
async def process_grading_date(message: Message, state: FSMContext):
    try:
        date = datetime.strptime(message.text.strip(), "%d-%m-%Y")
    except ValueError:
        await message.answer("Неверный формат даты. Введите в формате дд-мм-гггг.")
        return
    
    await state.update_data(grading_date=date.strftime("%d-%m-%Y"))
    keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="Быстрое выставление")],
            [KeyboardButton(text="Одиночное выставление")]
        ],
        resize_keyboard=True
    )
    await state.set_state(GradeForm.selecting_grading_type)
    await message.answer("Выберите тип выставления оценок:", reply_markup=keyboard)

@dp.message(GradeForm.selecting_grading_type, F.text.in_(["Быстрое выставление", "Одиночное выставление"]))
async def process_grading_type(message: Message, state: FSMContext):
    grading_type = message.text
    await state.update_data(grading_type=grading_type)
    data = await state.get_data()
    students = data.get("students", [])
    
    if grading_type == "Быстрое выставление":
        await state.set_state(GradeForm.selecting_student)
        student_id, full_name = students[0]
        await state.update_data(current_student=0)
        await message.answer(f"Студент #{1}: {full_name}\nВыберите оценку:", 
                           reply_markup=get_grades_keyboard())
    else:
        await state.set_state(GradeForm.selecting_student_single)
        keyboard = InlineKeyboardBuilder()
        for student_id, full_name in students:
            keyboard.add(InlineKeyboardButton(text=full_name, callback_data=f"grade_student_{student_id}"))
        keyboard.adjust(1)
        await message.answer("Выберите студента для выставления оценки:", reply_markup=keyboard.as_markup())

def get_grades_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="0"), KeyboardButton(text="1"), KeyboardButton(text="2")],
            [KeyboardButton(text="3"), KeyboardButton(text="4"), KeyboardButton(text="5")],
            [KeyboardButton(text="н"), KeyboardButton(text="Пропустить"), KeyboardButton(text="Отменить")]
        ],
        resize_keyboard=True
    )

@dp.message(GradeForm.selecting_student, F.text.in_(["0", "1", "2", "3", "4", "5", "н", "Пропустить", "Отменить"]))
async def process_grade(message: Message, state: FSMContext):
    data = await state.get_data()
    ktp_id = data.get("ktp_id")
    discipline_id = data.get("discipline_id")
    students = data.get("students", [])
    current_student = data.get("current_student", 0)
    teacher_id = message.from_user.id
    date = data.get("grading_date")
    
    if message.text == "Отменить":
        await state.clear()
        await message.answer("Выставление оценок отменено.", reply_markup=ReplyKeyboardRemove())
        await show_teacher_menu(message)
        return
    
    if message.text != "Пропустить":
        student_id, full_name = students[current_student]
        grade = -1 if message.text == "н" else int(message.text)
        
        cursor.execute("""
        INSERT INTO grades (student_id, teacher_id, discipline_id, ktp_id, grade, date)
        VALUES (?, ?, ?, ?, ?, ?)
        """, (student_id, teacher_id, discipline_id, ktp_id, grade, date))
        
        cursor.execute("SELECT tokens_per_attendance FROM teachers WHERE teacher_id = ?", (teacher_id,))
        tokens_per_attendance = cursor.fetchone()[0]
        if grade != -1:
            cursor.execute("""
            UPDATE student_teacher 
            SET tokens = tokens + ? 
            WHERE student_id = ? AND teacher_id = ?
            """, (tokens_per_attendance, student_id, teacher_id))
        
        # Получаем информацию для уведомления
        cursor.execute("SELECT name FROM disciplines WHERE discipline_id = ?", (discipline_id,))
        discipline_name = cursor.fetchone()[0]
        cursor.execute("SELECT description, homework FROM ktp WHERE ktp_id = ?", (ktp_id,))
        ktp_data = cursor.fetchone()
        ktp_description, homework = ktp_data
        cursor.execute("SELECT full_name FROM users WHERE user_id = ?", (teacher_id,))
        teacher_name = cursor.fetchone()[0]
        
        # Отправляем уведом Huntington уведомление студенту
        await notify_student(student_id, discipline_name, grade, date, ktp_description, teacher_name, homework=homework)
        
        conn.commit()
    
    current_student += 1
    if current_student >= len(students):
        await state.clear()
        await message.answer("Выставление оценок завершено.", reply_markup=ReplyKeyboardRemove())
        await show_teacher_menu(message)
        return
    
    await state.update_data(current_student=current_student)
    next_student_id, next_full_name = students[current_student]
    await message.answer(f"Студент #{current_student + 1}: {next_full_name}\nВыберите оценку:", 
                       reply_markup=get_grades_keyboard())

@dp.callback_query(F.data.startswith("grade_student_"))
async def select_student_for_grading(callback: types.CallbackQuery, state: FSMContext):
    student_id = int(callback.data.split("_")[2])
    data = await state.get_data()
    students = data.get("students", [])
    student_info = next((s for s in students if s[0] == student_id), None)
    
    if not student_info:
        await callback.message.answer("Студент не найден.")
        await callback.answer()
        return
    
    await state.update_data(selected_student_id=student_id, selected_student_name=student_info[1])
    await state.set_state(GradeForm.entering_grade)
    await callback.message.answer(f"Выберите оценку для {student_info[1]}:", reply_markup=get_grades_keyboard_single())
    await callback.answer()

def get_grades_keyboard_single():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="0"), KeyboardButton(text="1"), KeyboardButton(text="2")],
            [KeyboardButton(text="3"), KeyboardButton(text="4"), KeyboardButton(text="5")],
            [KeyboardButton(text="н")]
        ],
        resize_keyboard=True
    )

@dp.message(GradeForm.entering_grade, F.text.in_(["0", "1", "2", "3", "4", "5", "н", "Назад", "Завершить"]))
async def process_grade_single(message: Message, state: FSMContext):
    data = await state.get_data()
    ktp_id = data.get("ktp_id")
    discipline_id = data.get("discipline_id")
    teacher_id = message.from_user.id
    date = data.get("grading_date")
    students = data.get("students", [])
    student_id = data.get("selected_student_id")
    student_name = data.get("selected_student_name")
    
    if message.text == "Завершить":
        await state.clear()
        await message.answer("Выставление оценок завершено.", reply_markup=ReplyKeyboardRemove())
        await show_teacher_menu(message)
        return
    
    if message.text == "Назад":
        await state.set_state(GradeForm.selecting_student_single)
        keyboard = InlineKeyboardBuilder()
        for s_id, full_name in students:
            keyboard.add(InlineKeyboardButton(text=full_name, callback_data=f"grade_student_{s_id}"))
        keyboard.add(InlineKeyboardButton(text="Завершить", callback_data="finish_grading"))
        keyboard.adjust(1)
        await message.answer("Выберите студента для выставления оценки или завершите:", reply_markup=keyboard.as_markup())
        return
    
    grade = -1 if message.text == "н" else int(message.text)
    
    cursor.execute("SELECT grade_id FROM grades WHERE student_id = ? AND discipline_id = ? AND ktp_id = ?",
                  (student_id, discipline_id, ktp_id))
    existing_grade = cursor.fetchone()
    
    is_update = bool(existing_grade)
    if not existing_grade:
        cursor.execute("""
        INSERT INTO grades (student_id, teacher_id, discipline_id, ktp_id, grade, date)
        VALUES (?, ?, ?, ?, ?, ?)
        """, (student_id, teacher_id, discipline_id, ktp_id, grade, date))
        
        if grade != -1:
            cursor.execute("SELECT tokens_per_attendance FROM teachers WHERE teacher_id = ?", (teacher_id,))
            tokens_per_attendance = cursor.fetchone()[0]
            cursor.execute("""
            UPDATE student_teacher 
            SET tokens = tokens + ? 
            WHERE student_id = ? AND teacher_id = ?
            """, (tokens_per_attendance, student_id, teacher_id))
    else:
        cursor.execute("""
        UPDATE grades 
        SET grade = ?, date = ? 
        WHERE student_id = ? AND discipline_id = ? AND ktp_id = ?
        """, (grade, date, student_id, discipline_id, ktp_id))
    
    # Получаем информацию для уведомления
    cursor.execute("SELECT name FROM disciplines WHERE discipline_id = ?", (discipline_id,))
    discipline_name = cursor.fetchone()[0]
    cursor.execute("SELECT description, homework FROM ktp WHERE ktp_id = ?", (ktp_id,))
    ktp_data = cursor.fetchone()
    ktp_description, homework = ktp_data
    cursor.execute("SELECT full_name FROM users WHERE user_id = ?", (teacher_id,))
    teacher_name = cursor.fetchone()[0]
    
    # Отправляем уведомление студенту
    await notify_student(student_id, discipline_name, grade, date, ktp_description, teacher_name, homework=homework, is_update=is_update)
    
    conn.commit()
    
    await message.answer(f"Оценка {'изменена' if is_update else 'выставлена'} для {student_name}: {message.text}.")
    
    await state.set_state(GradeForm.selecting_student_single)
    keyboard = InlineKeyboardBuilder()
    for s_id, full_name in students:
        keyboard.add(InlineKeyboardButton(text=full_name, callback_data=f"grade_student_{s_id}"))
    keyboard.add(InlineKeyboardButton(text="Завершить", callback_data="finish_grading"))
    keyboard.adjust(1)
    await message.answer("Выберите студента для выставления оценки или завершите:", reply_markup=keyboard.as_markup())

@dp.callback_query(F.data == "finish_grading")
async def finish_grading(callback: types.CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.answer("Выставление оценок завершено.", reply_markup=ReplyKeyboardRemove())
    await show_teacher_menu(callback.message)
    await callback.answer()

@dp.message(F.text == "Редактировать оценку")
async def start_edit_grade(message: Message, state: FSMContext):
    if not is_teacher(message.from_user.id):
        await message.answer("Команда доступна только преподавателями.")
        return
    teacher_id = message.from_user.id
    disciplines = get_teacher_disciplines(teacher_id)
    
    if not disciplines:
        await message.answer("У вас нет дисциплин.")
        return
    
    keyboard = InlineKeyboardBuilder()
    for discipline in disciplines:
        discipline_id, name, group = discipline
        keyboard.add(InlineKeyboardButton(text=f"{name} ({group})", callback_data=f"edit_dis_{discipline_id}"))
    keyboard.adjust(1)
    await state.set_state(GradeForm.selecting_discipline_edit)
    await message.answer("Выберите дисциплину для редактирования оценок:", reply_markup=keyboard.as_markup())

@dp.callback_query(F.data.startswith("edit_dis_"))
async def select_discipline_for_edit_grade(callback: types.CallbackQuery, state: FSMContext):
    discipline_id = int(callback.data.split("_")[2])
    teacher_id = callback.from_user.id
    
    cursor.execute("SELECT name, group_name FROM disciplines WHERE discipline_id = ? AND teacher_id = ?", 
                  (discipline_id, teacher_id))
    discipline = cursor.fetchone()
    if not discipline:
        await callback.message.answer("Дисциплина не найдена.")
        await state.clear()
        await callback.answer()
        return
    
    await state.update_data(edit_discipline_id=discipline_id, edit_discipline_name=discipline[0], edit_group_name=discipline[1])
    today = datetime.now().strftime("%d-%m-%Y")
    keyboard = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text=today)]],
        resize_keyboard=True
    )
    await state.set_state(GradeForm.selecting_date_edit)
    await callback.message.answer(f"Введите дату для редактирования оценок (дд-мм-гггг, по умолчанию {today}):", reply_markup=keyboard)
    await callback.answer()

@dp.message(GradeForm.selecting_date_edit)
async def process_edit_grade_date(message: Message, state: FSMContext):
    try:
        date = datetime.strptime(message.text.strip(), "%d-%m-%Y")
    except ValueError:
        await message.answer("Неверный формат даты. Введите в формате дд-мм-гггг.")
        return
    
    teacher_id = message.from_user.id
    data = await state.get_data()
    discipline_id = data.get("edit_discipline_id")
    
    cursor.execute("""
    SELECT DISTINCT k.ktp_id, k.discipline_id, k.group_name, k.description 
    FROM ktp k
    JOIN grades g ON k.ktp_id = g.ktp_id
    WHERE k.teacher_id = ? AND g.date = ? AND k.discipline_id = ?
    """, (teacher_id, date.strftime("%d-%m-%Y"), discipline_id))
    ktps = cursor.fetchall()
    
    if not ktps:
        await message.answer(f"Нет КТП с оценками на {date.strftime('%d-%m-%Y')} для выбранной дисциплины.")
        await state.clear()
        return
    
    await state.update_data(edit_date=date.strftime("%d-%m-%Y"))
    keyboard = InlineKeyboardBuilder()
    for ktp_id, discipline_id, group_name, description in ktps:
        cursor.execute("SELECT name FROM disciplines WHERE discipline_id = ?", (discipline_id,))
        discipline_name = cursor.fetchone()[0]
        keyboard.add(InlineKeyboardButton(
            text=f"{discipline_name} ({group_name}, {description})",
            callback_data=f"edit_grade_ktp_{ktp_id}_{discipline_id}"
        ))
    keyboard.adjust(1)
    await state.set_state(GradeForm.selecting_ktp_edit)
    await message.answer("Выберите КТП для редактирования оценок:", reply_markup=keyboard.as_markup())

@dp.callback_query(F.data.startswith("edit_grade_ktp_"))
async def select_ktp_for_edit_grade(callback: types.CallbackQuery, state: FSMContext):
    ktp_id = int(callback.data.split("_")[3])
    discipline_id = int(callback.data.split("_")[4])
    teacher_id = callback.from_user.id
    
    cursor.execute("SELECT name, group_name FROM disciplines WHERE discipline_id = ?", (discipline_id,))
    discipline = cursor.fetchone()
    if not discipline:
        await callback.message.answer("Дисциплина не найдена.")
        await state.clear()
        await callback.answer()
        return
    
    discipline_name, group_name = discipline
    students = get_group_students(group_name, teacher_id)
    
    if not students:
        await callback.message.answer("В этой группе нет студентов.")
        await state.clear()
        await callback.answer()
        return
    
    await state.update_data(
        edit_ktp_id=ktp_id,
        edit_discipline_id=discipline_id,
        edit_discipline_name=discipline_name,
        edit_group_name=group_name,
        edit_students=students
    )
    
    keyboard = InlineKeyboardBuilder()
    for student_id, full_name in students:
        cursor.execute("SELECT grade FROM grades WHERE student_id = ? AND discipline_id = ? AND ktp_id = ?",
                     (student_id, discipline_id, ktp_id))
        grade = cursor.fetchone()
        grade_str = "н" if grade and grade[0] == -1 else str(grade[0]) if grade else "нет оценки"
        keyboard.add(InlineKeyboardButton(
            text=f"{full_name} ({grade_str})",
            callback_data=f"edit_grade_student_{student_id}"
        ))
    keyboard.add(InlineKeyboardButton(text="Отменить", callback_data="cancel_edit_grade"))
    keyboard.adjust(1)
    await state.set_state(GradeForm.selecting_student_edit)
    await callback.message.answer("Выберите студента для редактирования оценки:", reply_markup=keyboard.as_markup())
    await callback.answer()

@dp.callback_query(F.data.startswith("edit_grade_student_"))
async def select_student_for_edit_grade(callback: types.CallbackQuery, state: FSMContext):
    student_id = int(callback.data.split("_")[3])
    data = await state.get_data()
    students = data.get("edit_students", [])
    student_info = next((s for s in students if s[0] == student_id), None)
    
    if not student_info:
        await callback.message.answer("Студент не найден.")
        await callback.answer()
        return
    
    await state.update_data(edit_student_id=student_id, edit_student_name=student_info[1])
    await state.set_state(GradeForm.entering_grade_edit)
    await callback.message.answer(f"Введите новую оценку для {student_info[1]}:", reply_markup=get_grades_keyboard_single())
    await callback.answer()

@dp.callback_query(F.data == "cancel_edit_grade")
async def cancel_edit_grade(callback: types.CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.answer("Редактирование оценок отменено.", reply_markup=ReplyKeyboardRemove())
    await show_teacher_menu(callback.message)
    await callback.answer()

@dp.message(GradeForm.entering_grade_edit, F.text.in_(["0", "1", "2", "3", "4", "5", "н", "Назад", "Завершить"]))
async def process_edit_grade(message: Message, state: FSMContext):
    data = await state.get_data()
    ktp_id = data.get("edit_ktp_id")
    discipline_id = data.get("edit_discipline_id")
    discipline_name = data.get("edit_discipline_name")
    teacher_id = message.from_user.id
    date = data.get("edit_date")
    student_id = data.get("edit_student_id")
    student_name = data.get("edit_student_name")
    students = data.get("edit_students", [])
    
    if message.text == "Завершить":
        await state.clear()
        await message.answer("Редактирование оценок завершено.", reply_markup=ReplyKeyboardRemove())
        await show_teacher_menu(message)
        return
    
    if message.text == "Назад":
        await state.set_state(GradeForm.selecting_student_edit)
        keyboard = InlineKeyboardBuilder()
        for s_id, full_name in students:
            cursor.execute("SELECT grade FROM grades WHERE student_id = ? AND discipline_id = ? AND ktp_id = ?",
                         (s_id, discipline_id, ktp_id))
            grade = cursor.fetchone()
            grade_str = "н" if grade and grade[0] == -1 else str(grade[0]) if grade else "нет оценки"
            keyboard.add(InlineKeyboardButton(
                text=f"{full_name} ({grade_str})",
                callback_data=f"edit_grade_student_{s_id}"
            ))
        keyboard.add(InlineKeyboardButton(text="Отменить", callback_data="cancel_edit_grade"))
        keyboard.adjust(1)
        await message.answer("Выберите студента для редактирования оценки:", reply_markup=keyboard.as_markup())
        return
    
    grade = -1 if message.text == "н" else int(message.text)
    
    cursor.execute("SELECT grade_id, grade FROM grades WHERE student_id = ? AND discipline_id = ? AND ktp_id = ?",
                  (student_id, discipline_id, ktp_id))
    existing_grade = cursor.fetchone()
    
    is_update = bool(existing_grade)
    old_grade = existing_grade[1] if existing_grade else None
    
    if is_update:
        cursor.execute("""
        UPDATE grades 
        SET grade = ?, date = ? 
        WHERE student_id = ? AND discipline_id = ? AND ktp_id = ?
        """, (grade, date, student_id, discipline_id, ktp_id))
    else:
        cursor.execute("""
        INSERT INTO grades (student_id, teacher_id, discipline_id, ktp_id, grade, date)
        VALUES (?, ?, ?, ?, ?, ?)
        """, (student_id, teacher_id, discipline_id, ktp_id, grade, date))
    
    # Обновляем токены, если нужно
    if grade != -1 and (not is_update or (is_update and old_grade == -1)):
        cursor.execute("SELECT tokens_per_attendance FROM teachers WHERE teacher_id = ?", (teacher_id,))
        tokens_per_attendance = cursor.fetchone()[0]
        cursor.execute("""
        UPDATE student_teacher 
        SET tokens = tokens + ? 
        WHERE student_id = ? AND teacher_id = ?
        """, (tokens_per_attendance, student_id, teacher_id))
    elif grade == -1 and is_update and old_grade != -1:
        cursor.execute("SELECT tokens_per_attendance FROM teachers WHERE teacher_id = ?", (teacher_id,))
        tokens_per_attendance = cursor.fetchone()[0]
        cursor.execute("""
        UPDATE student_teacher 
        SET tokens = tokens - ? 
        WHERE student_id = ? AND teacher_id = ?
        """, (tokens_per_attendance, student_id, teacher_id))
    
    # Получаем информацию для уведомления
    cursor.execute("SELECT description, homework FROM ktp WHERE ktp_id = ?", (ktp_id,))
    ktp_data = cursor.fetchone()
    ktp_description, homework = ktp_data
    cursor.execute("SELECT full_name FROM users WHERE user_id = ?", (teacher_id,))
    teacher_name = cursor.fetchone()[0]
    
    # Отправляем уведомление студенту
    await notify_student(student_id, discipline_name, grade, date, ktp_description, teacher_name, homework=homework, is_update=is_update)
    
    conn.commit()
    
    await message.answer(f"Оценка {'изменена' if is_update else 'выставлена'} для {student_name}: {message.text}.")
    
    # Возвращаем к выбору студента
    await state.set_state(GradeForm.selecting_student_edit)
    keyboard = InlineKeyboardBuilder()
    for s_id, full_name in students:
        cursor.execute("SELECT grade FROM grades WHERE student_id = ? AND discipline_id = ? AND ktp_id = ?",
                     (s_id, discipline_id, ktp_id))
        grade = cursor.fetchone()
        grade_str = "н" if grade and grade[0] == -1 else str(grade[0]) if grade else "нет оценки"
        keyboard.add(InlineKeyboardButton(
            text=f"{full_name} ({grade_str})",
            callback_data=f"edit_grade_student_{s_id}"
        ))
    keyboard.add(InlineKeyboardButton(text="Отменить", callback_data="cancel_edit_grade"))
    keyboard.adjust(1)
    await message.answer("Выберите студента для редактирования оценки:", reply_markup=keyboard.as_markup())

@dp.message(F.text == "Мои студенты")
async def show_teacher_students(message: Message):
    teacher_id = message.from_user.id
    groups = get_teacher_groups(teacher_id)
    
    if not groups:
        await message.answer("У вас нет групп.")
        return
    
    keyboard = InlineKeyboardBuilder()
    for group in groups:
        keyboard.add(InlineKeyboardButton(text=group, callback_data=f"view_group_{group}"))
    keyboard.adjust(1)
    await message.answer("Выберите группу для просмотра студентов:", reply_markup=keyboard.as_markup())

@dp.callback_query(F.data.startswith("view_group_"))
async def view_group_students(callback: types.CallbackQuery):
    group_name = callback.data.split("_")[2]
    teacher_id = callback.from_user.id
    students = get_group_students(group_name, teacher_id)
    
    if not students:
        await callback.message.answer(f"В группе {group_name} нет студентов.")
        await callback.answer()
        return
    
    response = f"Студенты группы {group_name}:\n"
    for student_id, full_name in students:
        tokens = get_student_token_balance(student_id, teacher_id)
        response += f"- {full_name} ({tokens} жетончиков)\n"
    
    await callback.message.answer(response)
    await callback.answer()

@dp.message(F.text == "Создать ведомость")
async def create_gradebook_start(message: Message):
    if not is_teacher(message.from_user.id):
        await message.answer("Команда доступна только преподавателями.")
        return
    teacher_id = message.from_user.id
    disciplines = get_teacher_disciplines(teacher_id)
    
    if not disciplines:
        await message.answer("У вас нет дисциплин.")
        return
    
    keyboard = InlineKeyboardBuilder()
    for discipline in disciplines:
        discipline_id, name, group = discipline
        keyboard.add(InlineKeyboardButton(text=f"{name} ({group})", callback_data=f"book_dis_{discipline_id}"))
    keyboard.adjust(1)
    await message.answer("Выберите дисциплину для создания ведомости:", reply_markup=keyboard.as_markup())

@dp.callback_query(F.data.startswith("book_dis_"))
async def generate_gradebook(callback: types.CallbackQuery):
    discipline_id = int(callback.data.split("_")[2])
    teacher_id = callback.from_user.id
    
    try:
        filename = create_gradebook(teacher_id, discipline_id)
        await callback.message.answer_document(
            document=FSInputFile(path=filename),
            caption="Ведомость создана."
        )
        os.remove(filename)
    except Exception as e:
        logger.error(f"Error creating gradebook: {e}")
        await callback.message.answer("Ошибка при создания ведомости.")
    
    await callback.answer()

@dp.message(F.text == "Мои оценки")
async def show_student_grades(message: Message):
    student_id = message.from_user.id
    grades = get_student_grades(student_id)
    
    if not grades:
        await message.answer("У вас нет оценок за последние 30 дней.")
        return
    
    response = "Ваши оценки:\n"
    for grade, date, discipline, teacher, ktp, homework in grades:
        grade_str = "н" if grade == -1 else str(grade)
        response += f"- {date}: {discipline} ({ktp}) - {grade_str} (Преп.: {teacher})\n"
        if homework:
            response += f"  ДЗ: {homework}\n"
    
    await message.answer(response)

@dp.message(F.text == "Экспорт оценок")
async def export_student_grades(message: Message):
    student_id = message.from_user.id
    try:
        filename = create_grades_excel(student_id)
        document = FSInputFile(filename)
        await message.answer_document(document=document, 
                                    caption="Ваши оценки за последние 30 дней.")
        os.remove(filename)
    except Exception as e:
        logger.error(f"Error exporting grades: {e}")
        await message.answer("Ошибка при экспорте оценок.")

@dp.message(F.text == "Мои преподаватели")
async def show_student_teachers(message: Message):
    student_id = message.from_user.id
    teachers = get_student_teachers(student_id)
    
    if not teachers:
        await message.answer("У вас нет преподавателей.")
        return
    
    response = "Ваши преподаватели:\n"
    for teacher_id, full_name, tokens in teachers:
        response += f"- {full_name} ({tokens} жетончиков)\n"
    
    await message.answer(response)

@dp.message(F.text == "Мои жетончики")
async def show_student_tokens(message: Message):
    student_id = message.from_user.id
    teachers = get_student_teachers(student_id)
    
    if not teachers:
        await message.answer("У вас нет жетончиков.")
        return
    
    response = "Ваши жетончики:\n"
    for teacher_id, full_name, tokens in teachers:
        response += f"- {full_name}: {tokens} жетончиков\n"
    
    await message.answer(response)

@dp.message(F.text == "Магазин наград")
async def show_rewards_shop(message: Message):
    student_id = message.from_user.id
    rewards = get_student_rewards(student_id)
    
    if not rewards:
        await message.answer("В магазине нет наград.")
        return
    
    response = "Магазин наград:\n"
    keyboard = InlineKeyboardBuilder()
    for reward_id, name, description, price, teacher_name, discipline_name, teacher_id in rewards:
        tokens = get_student_token_balance(student_id, teacher_id)
        status = " (доступно)" if tokens >= price else " (недостаточно жетончиков)"
        response += f"- {name} ({discipline_name}): {description}, {price} жетончиков{status} (Преп.: {teacher_name})\n"
        if tokens >= price:
            keyboard.add(InlineKeyboardButton(text=f"Купить: {name}", callback_data=f"buy_reward_{reward_id}"))
    keyboard.adjust(1)
    await message.answer(response, reply_markup=keyboard.as_markup())

@dp.callback_query(F.data.startswith("buy_reward_"))
async def buy_reward(callback: types.CallbackQuery):
    reward_id = int(callback.data.split("_")[2])
    student_id = callback.from_user.id
    
    cursor.execute("""
    SELECT r.price, r.name, r.teacher_id, d.name
    FROM rewards r
    JOIN disciplines d ON r.discipline_id = d.discipline_id
    WHERE r.reward_id = ?
    """, (reward_id,))
    reward = cursor.fetchone()
    
    if not reward:
        await callback.message.answer("Награда не найдена.")
        await callback.answer()
        return
    
    price, reward_name, teacher_id, discipline_name = reward
    tokens = get_student_token_balance(student_id, teacher_id)
    
    if tokens < price:
        await callback.message.answer("Недостаточно жетончиков для покупки.")
        await callback.answer()
        return
    
    cursor.execute("""
    UPDATE student_teacher 
    SET tokens = tokens - ? 
    WHERE student_id = ? AND teacher_id = ?
    """, (price, student_id, teacher_id))
    
    cursor.execute("""
    INSERT INTO purchased_rewards (student_id, reward_id, teacher_id, purchase_date)
    VALUES (?, ?, ?, ?)
    """, (student_id, reward_id, teacher_id, datetime.now().strftime("%d-%m-%Y")))
    
    conn.commit()
    
    cursor.execute("SELECT full_name FROM users WHERE user_id = ?", (teacher_id,))
    teacher_name = cursor.fetchone()[0]
    cursor.execute("SELECT full_name FROM users WHERE user_id = ?", (student_id,))
    student_name = cursor.fetchone()[0]
    
    await callback.message.answer(f"Вы купили награду '{reward_name}' за {price} жетончиков!")
    await bot.send_message(
        chat_id=teacher_id,
        text=f"Студент {student_name} купил награду '{reward_name}' ({discipline_name}) за {price} жетончиков."
    )
    await callback.answer()

@dp.message(F.text == "Мои практики")
async def check_student_practices(message: Message, state: FSMContext):
    if not is_student(message.from_user.id):
        await message.answer("Команда доступна только студентам.")
        return
    
    student_id = message.from_user.id
    try:
        # Получаем дисциплины, связанные со студентом через группу
        cursor.execute("""
        SELECT d.discipline_id, d.name, d.required_practices
        FROM disciplines d
        JOIN student_teacher st ON d.teacher_id = st.teacher_id
        WHERE st.student_id = ? AND d.group_name = (SELECT group_name FROM students WHERE student_id = ?)
        """, (student_id, student_id))
        disciplines = cursor.fetchall()
        
        if not disciplines:
            await message.answer("У вас нет дисциплин.")
            return
        
        response = "Ваши практики:\n\n"
        for discipline_id, name, required_practices in disciplines:
            # Подсчитываем сданные практики
            cursor.execute("""
            SELECT COUNT(DISTINCT k.practice_number)
            FROM grades g
            JOIN ktp k ON g.ktp_id = k.ktp_id
            WHERE g.student_id = ? AND g.discipline_id = ? AND k.type = 'practice'
            AND g.grade >= 1 AND g.grade != -1
            """, (student_id, discipline_id))
            completed = cursor.fetchone()[0]
            
            required = required_practices if required_practices is not None else 0
            response += f"Дисциплина: {name}\n"
            response += f"Сдано практик: {completed} из {required}\n"
            response += f"Статус: {'Все практики сданы' if completed >= required and required > 0 else 'Есть несданные практики'}\n\n"
        
        await message.answer(response)
    except sqlite3.OperationalError as e:
        logger.error(f"Ошибка при проверке практик студента {student_id}: {e}")
        await message.answer("Ошибка при получении данных о практиках.")
    except Exception as e:
        logger.error(f"Неизвестная ошибка при проверке практик: {e}")
        await message.answer("Произошла ошибка. Попробуйте позже.")

@dp.message(F.text == "Загрузить список преподавателей")
async def upload_teachers_start(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("Команда доступна только администраторам.")
        return
    await message.answer("Загрузите Excel файл со списком преподавателей (колонка: ФИО преподавателя).")

@dp.message(F.text == "Сгенерировать токены преподавателей")
async def generate_teacher_tokens(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("Команда доступна только администраторам.")
        return
    
    cursor.execute("SELECT teacher_id, full_name FROM teachers")
    teachers = cursor.fetchall()
    
    if not teachers:
        await message.answer("Нет преподавателей.")
        return
    
    tokens = []
    for teacher_id, full_name in teachers:
        cursor.execute("SELECT token FROM teachers WHERE teacher_id = ?", (teacher_id,))
        token = cursor.fetchone()[0]
        if not token:
            token = generate_token()
            cursor.execute("UPDATE teachers SET token = ? WHERE teacher_id = ?", (token, teacher_id))
        tokens.append((full_name, token))
    
    conn.commit()
    
    df = pd.DataFrame(tokens, columns=["ФИО преподавателя", "Токен"])
    filename = f"teacher_tokens_{datetime.now().strftime('%Y%m%d')}.xlsx"
    df.to_excel(filename, index=False)
    
    await message.answer_document(
        document=FSInputFile(path=filename),
        caption="Токены сгенерированы."
    )
    os.remove(filename)

@dp.message(F.text == "Сгенерировать токены преподавателей")
async def generate_teacher_tokens(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("Команда доступна только администратору.")
        return
    teachers = get_admin_teachers(message.from_user.id)
    
    if not teachers:
        await message.answer("Нет преподавателей для генерации токенов.")
        return
    
    response = "Токены преподавателей:\n"
    for teacher_id, full_name in teachers:
        cursor.execute("SELECT token FROM teachers WHERE teacher_id = ?", (teacher_id,))
        token = cursor.fetchone()
        if not token:
            new_token = generate_token()
            cursor.execute("UPDATE teachers SET token = ? WHERE teacher_id = ?", (new_token, teacher_id))
            conn.commit()
            response += f"- {full_name}: {new_token}\n"
        else:
            response += f"- {full_name}: {token[0]}\n"
    
    await message.answer(response)

@dp.message(F.text == "Просмотреть список преподавателей")
async def list_teachers(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("Команда доступна только администратору.")
        return
    teachers = get_admin_teachers(message.from_user.id)
    
    if not teachers:
        await message.answer("Нет преподавателей.")
        return
    
    response = "Список преподавателей:\n"
    for teacher_id, full_name in teachers:
        response += f"- {full_name}\n"
    
    await message.answer(response)

@dp.message(F.text == "Просмотреть список студентов")
async def list_students(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("Команда доступна только администратору.")
        return
    students = get_admin_students(message.from_user.id)
    
    if not students:
        await message.answer("Нет студентов.")
        return
    
    response = "Список студентов:\n"
    for student_id, full_name, group_name in students:
        response += f"- {full_name} ({group_name})\n"
    
    await message.answer(response)

@dp.message(F.text == "Шаблон преподавателей")
async def send_teachers_template(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("Команда доступна только администратору.")
        return
    filename = create_teachers_template()
    document = FSInputFile(filename)
    await message.answer_document(document=document, caption="Шаблон для преподавателей.")
    os.remove(filename)

@dp.message(F.text == "Шаблон студентов")
async def send_students_template(message: Message):
    if not is_teacher(message.from_user.id):
        await message.answer("Команда доступна только преподавателям.")
        return
    
    filename = create_students_template()
    await message.answer_document(
        document=FSInputFile(path=filename),
        caption="Шаблон для загрузки студентов."
    )
    os.remove(filename)

@dp.message(F.text == "Отправить новость")
async def admin_start_news(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await message.answer("Команда доступна только администраторам.")
        return
    
    keyboard = InlineKeyboardBuilder()
    keyboard.add(InlineKeyboardButton(text="Студенты", callback_data="news_students"))
    keyboard.add(InlineKeyboardButton(text="Преподаватели", callback_data="news_teachers"))
    keyboard.add(InlineKeyboardButton(text="Все", callback_data="news_all"))
    keyboard.adjust(1)
    
    await state.set_state(AdminForm.waiting_for_recipient)
    await message.answer("Выберите получателей новости:", reply_markup=keyboard.as_markup())

@dp.callback_query(AdminForm.waiting_for_recipient, lambda c: c.data.startswith("news_"))
async def admin_process_recipient(callback: types.CallbackQuery, state: FSMContext):
    recipient = callback.data.split("_")[1]
    await state.update_data(recipient=recipient)
    await state.set_state(AdminForm.waiting_for_news_message)
    await callback.message.answer("Введите текст новости:")
    await callback.answer()

@dp.message(AdminForm.waiting_for_news_message)
async def admin_send_news(message: Message, state: FSMContext):
    try:
        data = await state.get_data()
        recipient = data.get("recipient")
        news_message = message.text.strip()
        
        if not news_message:
            await message.answer("Сообщение не может быть пустым. Попробуйте снова.")
            return
        
        # Жёстко прописываем имя администратора
        admin_name = "администратора"
        
        cursor.execute("SELECT user_id FROM users WHERE role = ? OR ? = 'all'", 
                      ('student' if recipient == 'students' else 'teacher' if recipient == 'teachers' else '', recipient))
        users = cursor.fetchall()
        
        if not users:
            await message.answer("Нет пользователей для отправки новости.")
            await state.clear()
            return
        
        sent_count = 0
        for user_id, in users:
            try:
                await bot.send_message(user_id, f"📢 Новость от {admin_name}: {news_message}")
                sent_count += 1
            except Exception as e:
                logger.error(f"Ошибка отправки новости пользователю {user_id}: {e}")
        
        await message.answer(f"Новость отправлена {sent_count} пользователям.")
        await state.clear()
    except sqlite3.OperationalError as e:
        logger.error(f"Ошибка при отправке новости: {e}")
        await message.answer("Ошибка при отправке новости.")
        await state.clear()
    except Exception as e:
        logger.error(f"Неизвестная ошибка при отправке новости: {e}")
        await message.answer("Произошла ошибка. Попробуйте позже.")
        await state.clear()

@dp.message(F.text == "Отправить новость группе")
async def teacher_start_news(message: Message, state: FSMContext):
    if not is_teacher(message.from_user.id):
        await message.answer("Команда доступна только преподавателям.")
        return
    
    teacher_id = message.from_user.id
    groups = get_teacher_groups(teacher_id)
    
    if not groups:
        await message.answer("У вас нет групп.")
        await state.clear()
        return
    
    keyboard = InlineKeyboardBuilder()
    for group in groups:
        keyboard.add(InlineKeyboardButton(text=group, callback_data=f"news_group_{group}"))
    keyboard.adjust(1)
    
    await state.set_state(TeacherForm.waiting_for_group_news)
    await message.answer("Выберите группу для отправки новости:", reply_markup=keyboard.as_markup())

@dp.callback_query(TeacherForm.waiting_for_group_news, lambda c: c.data.startswith("news_group_"))
async def teacher_process_group_news(callback: types.CallbackQuery, state: FSMContext):
    group_name = callback.data.split("_")[2]
    await state.update_data(group_name=group_name)
    await state.set_state(TeacherForm.waiting_for_news_message)
    await callback.message.answer("Введите текст новости:")
    await callback.answer()

@dp.message(TeacherForm.waiting_for_news_message)
async def teacher_send_news(message: Message, state: FSMContext):
    try:
        data = await state.get_data()
        group_name = data.get("group_name")
        news_message = message.text.strip()
        
        if not news_message:
            await message.answer("Сообщение не может быть пустым. Попробуйте снова.")
            return
        
        # Получаем имя преподавателя
        cursor.execute("SELECT full_name FROM users WHERE user_id = ?", (message.from_user.id,))
        teacher_name = cursor.fetchone()
        teacher_name = teacher_name[0] if teacher_name else "Преподаватель"
        
        cursor.execute("SELECT student_id FROM students WHERE group_name = ?", (group_name,))
        students = cursor.fetchall()
        
        if not students:
            await message.answer("В этой группе нет студентов.")
            await state.clear()
            return
        
        sent_count = 0
        for student_id, in students:
            try:
                await bot.send_message(student_id, f"📢 Новость от {teacher_name}: {news_message}")
                sent_count += 1
            except Exception as e:
                logger.error(f"Ошибка отправки новости студенту {student_id}: {e}")
        
        await message.answer(f"Новость отправлена {sent_count} студентам группы '{group_name}'.")
        await state.clear()
    except sqlite3.OperationalError as e:
        logger.error(f"Ошибка при отправке новости группе {group_name}: {e}")
        await message.answer("Ошибка при отправке новости.")
        await state.clear()
    except Exception as e:
        logger.error(f"Неизвестная ошибка при отправке новости: {e}")
        await message.answer("Произошла ошибка. Попробуйте позже.")
        await state.clear()

async def main():
    try:
        await dp.start_polling(bot)
    finally:
        conn.close()

if __name__ == "__main__":
    import asyncio
    asyncio.run(main())
